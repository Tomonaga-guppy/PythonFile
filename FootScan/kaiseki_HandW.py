
import matplotlib.pyplot as plt
import cv2
import numpy as np
import openpyxl as px
import os

#%matplotlib inline

# 接地面取得
def area(im,pho, ikiti):
    numbe = pho
    numbe = numbe[:3] # 先頭から二文字抽出
    numbe2 = numbe[1:2] # 二文字目
    numbe3 = numbe[2:3] # 三文字目
    if numbe2 == '_':
        numbe_1 = numbe[:1] # 一文字目まで
        numbe = numbe_1
    elif numbe3 == '_':
        numbe_2 = numbe[:2] # 二文字目まで
        numbe = numbe_2
    else :
        numbe = numbe

    name = numbe
    global rera
    global rera1
    global rera2

    im = cv2.flip(im, 1) # 反転
    x = 0

    while x <= 2:
        x += 1
        
        if x == 1:
            img = im[95:3445,55:1235]
        elif x == 2: 
            img = im[95:3445,1300:2480]
        elif x == 3:
            break
        

        wrs = "/Users/BRLAB/others/img6/sole" + str(x) + ".jpg"
        cv2.imwrite(wrs, img)
        imgcope = cv2.imread(wrs)

        # 足裏周りのノイズ除去_____________________________________________________
        # 足裏型凸包の外側を真っ黒にし、そのマスクを取得、元の画像に合成
        (takasa, haba, nanka) = imgcope.shape
        chushin_y = int(takasa/2)
        chushin_x = int(haba/2)
        # print("中心は、", chushin_x,chushin_y)
        '''
        #旧型のマスク
        imgcope_HSV = cv2.cvtColor(imgcope, cv2.COLOR_BGR2HSV)# HSV 色相(Hue)彩度(Saturation)明度(Value・Brightness)
        imgcope_HSV = cv2.GaussianBlur(imgcope_HSV, (5, 5), 0)
        imgco_H, imgco_S, imgco_V = cv2.split(imgcope_HSV)
        _thre, mask_a = cv2.threshold(imgco_V, 65, 255, cv2.THRESH_BINARY) # 二値化
        mask_b = cv2.dilate(mask_a,np.ones((10,10),np.uint8),iterations = 1)
        '''
        # 新なマスク用足裏取得
        img_Lab = cv2.cvtColor(img, cv2.COLOR_BGR2Lab)
        img_Lab_L, img_Lab_a, img_Lab_b = cv2.split(img_Lab)
        pix = 3001
        img_L_r = cv2.adaptiveThreshold(img_Lab_L, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, pix, 0)
        img_L_r = cv2.dilate(img_L_r,np.ones((10,10),np.uint8),iterations = 1)
        kernel = np.ones((10,10),np.uint8)
        img_L_r = cv2.morphologyEx(img_L_r, cv2.MORPH_OPEN, kernel)
        img_L_r = cv2.dilate(img_L_r,np.ones((10,10),np.uint8),iterations = 1)
        
        #adaptiveThreshold メソッドを用いれば、あるピクセルを 2 値化したい場合、
        # そのピクセルを中心とする n×n ピクセルのデータを用いて、閾値計算を行い、2 値化を行う。
        # adaptiveThreshold は次のように、6 つの引数を使って指定する。
        #cv2.adaptiveThreshold(img, maxValue, adaptiveMethod, thresholdType, blockSize, C)
        #img は、入力画像で、maxValue は閾値を満たすピクセルに与える値となる。
        # 続いて、adaptiveMethod は適応閾値処理の種類を指定して平均値 ADAPTIVE_THRESH_MEAN_C あるいは
        # 標準化された平均値 ADAPTIVE_THRESH_GAUSSIAN_C のどちらかを指定する。
        # また、thresholdType には、上で示した 5 つの閾値処理の種類を指定する。
        # blockSize は、近傍ピクセルのサイズを奇数で指定する。
        # また、C には定数を指定する。閾値処理がすべて終えたときに得られた閾値から最終的に定数 C の値を引いて、最終結果として出力される。
        wr_m = "/Users/BRLAB/others/img6/z_mask_b" + str(x) + ".jpg"
        cv2.imwrite(wr_m, img_L_r)
        mask_b = cv2.imread(wr_m)

        mask_b = cv2.cvtColor(mask_b, cv2.COLOR_BGR2GRAY)
        # 凸包（hull） ゲット！！！
        # 作成できる外接矩形を全て試し、足裏形状を囲うもののみ選択
        contours, hierarchy = cv2.findContours(mask_b, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE) 
        mask_b = cv2.imread(wr_m)
        for i in range(len(contours)):
            # 最も大きい輪郭に外接矩形を作成し、画像に重ね書き
            rect = cv2.minAreaRect(contours[i])
            ((xx,yy),(w, h),th) = cv2.minAreaRect(contours[i])

            if w*h > 1000000 : # >1000000 は大きい輪郭のみを残すため(330*330 ぐらい)
                #     points = cv2.boxPoints(box)、
                # 引数 ... box : ((左上X座標, 左上Y座標)，(幅, 高さ)，回転角)
                # 戻り値 ... points : 4隅の頂点
                box = cv2.boxPoints(rect) # 外接矩形は描かないことにしたよ。
                box = np.int0(box)
                hull = cv2.convexHull(contours[i]) #  凸包
                mask_b = cv2.drawContours(mask_b,[hull],0,(0,0,255),3) # 凸包

        for a in range(len(hull)): # 凸包（hull）を拡大させる、中心から遠いほど伸びは大きくなる（足先やかかと）
            [p1] =  hull[a]
            x1_sa = p1[0] - chushin_x 
            y1_sa = p1[1] - chushin_y
            x1 = p1[0] + (x1_sa)/20
            y1 = p1[1] + (y1_sa)/23
            if x1 < 0:
                x1 = 0
            elif x1 > haba:
                x1 = haba
            if y1 < 0:
                y1 = 0
            elif y1 > takasa:
                y1 = takasa
            p1[0] = x1
            p1[1] = y1
            hull[a] = [p1]
        #mask_c = cv2.drawContours(imgcope,[hull],0,(0,0,255),3) # 凸包
        #wr_c = "/Users/BRLAB/others/img3/z_mask_c" + str(x) + ".jpg"
        #cv2.imwrite(wr_c, mask_c)
        mask_d = cv2.fillConvexPoly(mask_b, hull, (0,0,255))
        #wr = "/Users/BRLAB/others/img3/z_mask_d" + str(x) + ".jpg"
        #cv2.imwrite(wr, mask_d)
        mask_d = cv2.cvtColor(mask_d, cv2.COLOR_BGR2HSV)
        mask_H, mask_S, mask_V = cv2.split(mask_d)
        _thre, mask_e = cv2.threshold(mask_S, 100, 255, cv2.THRESH_BINARY_INV)
        wr_e = "/Users/BRLAB/others/img6/z_mask_e" + str(x) + ".jpg" # 周り白でマスク部分が黒の保存してるマスク
        cv2.imwrite(wr_e, mask_e)
        mask_e = cv2.imread(wr_e)
        # マスクを重ねてノイズ除去____________________________________________________________
        rows,cols,channels = mask_e.shape
        roi = imgcope[0:rows, 0:cols ]

        img2gray = cv2.cvtColor(mask_e,cv2.COLOR_BGR2GRAY)
        ret, mask = cv2.threshold(img2gray, 120, 255, cv2.THRESH_BINARY)
        mask_inv = cv2.bitwise_not(mask)
        img1_bg = cv2.bitwise_and(roi,roi,mask = mask_inv)
        wr_im = "/Users/BRLAB/others/img6/z_mask_im"+ str(x) + ".jpg"
        cv2.imwrite(wr_im, img1_bg)
        img = cv2.imread(wr_im)



        # HSV の V 取得_______________________________________________________
        img_HSV = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)# HSV 色相(Hue)彩度(Saturation)明度(Value・Brightness)
        img_HSV = cv2.GaussianBlur(img_HSV, (3, 3), 0)

        img_H, img_S, img_V = cv2.split(img_HSV) # ＢＧＲの分割(今回はHSV)


        # 65 (足裏)__________________________________________________________________
        i = 65
        _thre, img_mk = cv2.threshold(img_V, i, 255, cv2.THRESH_BINARY) # 二値化
        wr_65 = "/Users/BRLAB/others/img6/zz" + str(i) + str(x) + ".jpg"
        cv2.imwrite(wr_65, img_mk)# そのままでは色が塗れない。一度保存してからインプットした画像なら色が塗れる。（画像空間の問題）
        img_u= cv2.imread(wr_65)# img_V(明度のみ取得した画像)を一度保存してからインプットすることにより、画像空間をBGRで取れて、色を塗れるようになる。

        # 塗り替え
        img_uc = cv2.cvtColor(img_u, cv2.COLOR_BGR2HSV)

        # H(色相) は0～180、0,180(赤),75(緑っぽい),100(薄青),179(赤),255(緑っぽい)=255-180=75 !,
        h=img_uc[:,:,(0)]
        h=np.where((h<10) & (h>=0),120,h)
        img_uc[:,:,(0)]=h
        # S(彩度) は0～255、256=0
        s=img_uc[:,:,(1)]
        s=np.where((s<10) & (s>=0),255,s)
        img_uc[:,:,(1)]=s
        # V(明度) は0～255、256=0
        v=img_uc[:,:,(2)]
        v =np.where((v<=255) & (v>=100),255,v)
        img_uc[:,:,(2)]=v

        img_65=cv2.cvtColor(img_uc, cv2.COLOR_HSV2BGR)
        cv2.imwrite(wr_65, img_65)
        area1 = int(cv2.countNonZero(img_mk))
        #print("足裏：",area1,"/",img_mk.size)


        # 105（接地面）_____________________________________________________________________
        img_Lab = cv2.cvtColor(img, cv2.COLOR_BGR2Lab)
        #cc=img_Lab[1000,500,:]
        #print(cc)
        img_Lab_L, img_Lab_a, img_Lab_b = cv2.split(img_Lab)


        #L*a*b* 色空間は、ヒトの感覚に合わせて考案された色空間である。
        # L* は明度で、0 ≤ L* ≤ 100 の値を取りうる。a* および b* は色味を表している。
        # a* は緑色と赤色を表し、a* が正の大きな値になるほど赤みが強くなり、負の大きな値になるほど緑みが強くなり、また 0 であれば無彩色となる。
        # また、b* は青色と黄色を表し、a* が正の大きな値になるほど黄色みが強くなり、負の大きな値になるほど青みが強くなり、また 0 であれば無彩色となる。
        # a* および b* の最小値と最大値は、明度 L の値によって異なる
        # L*a*b* 色空間において、L* は 0 ≤ L* ≤ 100、a* および b* はマイナスからプラスまでの値を取りうる。
        # OpenCV においては、0 ≤ L ≤ 100、-127 ≤ a ≤ 127 および -127 ≤ b ≤ 127 で定義されている。
        # ただし、OpenCV の cv2.cvtColor メソッドを用いて変換された場合は、0 ≤ L ≤ 255、0 ≤ a ≤ 255 および 0 ≤ b ≤ 255 となる。
        # OpenCV で L*a*b* 色空間を取り扱うときは、ここを注意する必要がある。

        _thre, img_mk2 = cv2.threshold(img_Lab_b, ikiti, 255, cv2.THRESH_BINARY) # 二値化
        wr = "/Users/BRLAB/others/img6/zz" + str(ikiti) + str(x) + ".jpg"
        cv2.imwrite(wr, img_mk2)
        img_u= cv2.imread(wr)

        # 塗り替え
        img_uc = cv2.cvtColor(img_u, cv2.COLOR_BGR2HSV)

        # H(色相) は0～180、
        h=img_uc[:,:,(0)]
        h=np.where((h<10) & (h>=0),87,h)
        img_uc[:,:,(0)]=h
        # S(彩度) は0～255、256=0
        s=img_uc[:,:,(1)]
        s=np.where((s<10) & (s>=0),150,s)
        img_uc[:,:,(1)]=s
        # V(明度) は0～255、256=0
        v=img_uc[:,:,(2)]
        v =np.where((v<=255) & (v>=100),255,v)
        img_uc[:,:,(2)]=v

        img_105 =cv2.cvtColor(img_uc, cv2.COLOR_HSV2BGR)
        cv2.imwrite(wr, img_105)
        area2 = int(cv2.countNonZero(img_mk2))
        #print("接地面：",area2,"/",img_mk.size)
        ratio = area2 / area1 *100
        ra1 = format(ratio,'.1f')
        #print("接地面/足裏：",area2,"/",area1,"...",ra1,"%")

        # 重ねる______________________________________________________

        rows,cols,channels = img_105.shape
        roi = img_65[0:rows, 0:cols ]

        img2gray = cv2.cvtColor(img_105,cv2.COLOR_BGR2GRAY)
        ret, mask = cv2.threshold(img2gray, 120, 255, cv2.THRESH_BINARY)
        mask_inv = cv2.bitwise_not(mask)

        img1_bg = cv2.bitwise_and(roi,roi,mask = mask_inv)
        img2_fg = cv2.bitwise_and(img_105,img_105,mask = mask)
        img2_fg_src = cv2.cvtColor(img2_fg, cv2.COLOR_BGR2RGB)
        plt.imshow(img2_fg_src)

        dst = cv2.add(img1_bg,img2_fg)
        img_65[0:rows, 0:cols ] = dst
        i = 65161
        wr = "/Users/BRLAB/others/img6/zz" + str(i) + str(x) + ".jpg"
        cv2.imwrite(wr, img_65)

        src3 = cv2.cvtColor(img_65, cv2.COLOR_BGR2RGB)
        #plt.imshow(src3)
        #plt.show()
        # ________________________________________________________________\
        
        #tx1 = 'blue : ' + str(area1) + '  lightblue : ' + str(area2)
        tx1 = "Contact Area"
        ra = format(ratio,'.5f') # ratioの桁数指定
        tx2 = 'ratio :' + str(ra)
        if x == 1:
            rera1 = ratio
        elif x == 2:
            rera2 = ratio
        ra1 = format(ratio,'.1f')
        tx3 = 'lightblue / blue = ' + str(ra1) + '%' 

        cv2.putText(img_65,tx1, (100, 100), cv2.FONT_HERSHEY_PLAIN,3, (255, 255, 255),4, cv2.LINE_8)
        #cv2.putText(img_65,tx2, (100, 200), cv2.FONT_HERSHEY_PLAIN,3, (255, 255, 255),4, cv2.LINE_8)
        cv2.putText(img_65,tx3, (100, 200), cv2.FONT_HERSHEY_PLAIN,3, (255, 255, 255),4, cv2.LINE_8)
        #画像データ、テキスト、テキスト右下座標、フォント、文字の縮尺、色（青, 緑, 赤）、文字の太さ、タイプライン（アルゴリズム）、
        
        cv2.imwrite(wr, img_65)


    #繋ぐ___________________________________________________________

    i = 65161
    x = 1
    wr1 = "/Users/BRLAB/others/img6/zz" + str(i) + str(x) + ".jpg"
    i01 = cv2.imread(wr1)
    os.remove(wr1)
    wr_im = "/Users/BRLAB/others/img6/z_mask_im"+ str(x) + ".jpg" # sole & mask
    os.remove(wr_im)
    wr_m = "/Users/BRLAB/others/img6/z_mask_b" + str(x) + ".jpg" # original mask
    os.remove(wr_m)
    wr_65 = "/Users/BRLAB/others/img6/zz65" + str(x) + ".jpg"
    os.remove(wr_65)

    x = 2
    wr2 = "/Users/BRLAB/others/img6/zz" + str(i) + str(x) + ".jpg"
    i02 = cv2.imread(wr2)
    os.remove(wr2)
    wr_im = "/Users/BRLAB/others/img6/z_mask_im"+ str(x) + ".jpg"
    os.remove(wr_im)
    wr_m = "/Users/BRLAB/others/img6/z_mask_b" + str(x) + ".jpg"
    os.remove(wr_m)
    wr_65 = "/Users/BRLAB/others/img6/zz65" + str(x) + ".jpg"
    os.remove(wr_65)

    im_h = cv2.hconcat([i01, i02])
    res = "/Users/BRLAB/others/img6/result_3area_" + str(name) + ".jpg"
    cv2.imwrite(res, im_h)

    rera = (rera1 + rera2)/2

    return rera, name  # 画像と接地面比率（％）


# 片足の「しわ」と「RGB」取得
def edge_rgb(name, ikiti):
    x = 0
    while x <= 2:
        x += 1
        if x == 3:
            break
        wrs = "/Users/BRLAB/others/img6/sole" + str(x) + ".jpg"
        sole_one = cv2.imread(wrs)

        img_Lab = cv2.cvtColor(sole_one, cv2.COLOR_BGR2Lab)
        img_Lab_L, img_Lab_a, img_Lab_b = cv2.split(img_Lab)
        res_L = "/Users/BRLAB/others/img6/Lab_L" + str(name) +"_"+ str(x) +".jpg"
        cv2.imwrite(res_L, img_Lab_L)
        pix = 19
        img_L_r = cv2.adaptiveThreshold(img_Lab_L, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, pix, 0)
        #adaptiveThreshold メソッドを用いれば、あるピクセルを 2 値化したい場合、
        # そのピクセルを中心とする n×n ピクセルのデータを用いて、閾値計算を行い、2 値化を行う。
        # adaptiveThreshold は次のように、6 つの引数を使って指定する。
        #cv2.adaptiveThreshold(img, maxValue, adaptiveMethod, thresholdType, blockSize, C)
        #img は、入力画像で、maxValue は閾値を満たすピクセルに与える値となる。
        # 続いて、adaptiveMethod は適応閾値処理の種類を指定して平均値 ADAPTIVE_THRESH_MEAN_C あるいは
        # 標準化された平均値 ADAPTIVE_THRESH_GAUSSIAN_C のどちらかを指定する。
        # また、thresholdType には、上で示した 5 つの閾値処理の種類を指定する。
        # blockSize は、近傍ピクセルのサイズを奇数で指定する。
        # また、C には定数を指定する。閾値処理がすべて終えたときに得られた閾値から最終的に定数 C の値を引いて、最終結果として出力される。

        #ガウシアンフィルタ (Gaussian filter) 
        img_L_r  = cv2.GaussianBlur(img_L_r , (11,11), 1)
        #ガウス分布 (正規分布)で重み付けしたカーネルを使用するフィルタ
        res_d = "/Users/BRLAB/others/img6/dst" + str(name) +"_"+ str(x) +".jpg"
        cv2.imwrite(res_d, img_L_r)
        img_at_b = cv2.imread(res_d)
        img_at_bg = cv2.cvtColor(img_at_b, cv2.COLOR_BGR2GRAY)
        _thre, img_bz = cv2.threshold(img_at_bg, 180, 255, cv2.THRESH_BINARY) 
        kernel = np.ones((3,3),np.uint8)
        img_bz = cv2.morphologyEx(img_bz, cv2.MORPH_CLOSE, kernel)
        dst = cv2.morphologyEx(img_bz, cv2.MORPH_OPEN, kernel)
        res_d = "/Users/BRLAB/others/img6/dst" + str(name) +"_"+ str(x) +".jpg"
        cv2.imwrite(res_d, dst)
        # しわ画像ゲット
        # **************************************************************************************************

        # 接地面にマス目を与えて、200*200 がすべて接地面ならedge, RGBデータ取得し、硬度, 水分量計算
        co_a = "/Users/BRLAB/others/img6/zz" + str(ikiti) + str(x) + ".jpg"
        contact_area = cv2.imread(co_a)
        # 凸包（hull） ゲット！！！
        # 作成できる外接矩形を全て試し、足裏形状を囲うもののみ選択
        contact_area = cv2.cvtColor(contact_area, cv2.COLOR_BGR2GRAY)
        contours, hierarchy = cv2.findContours(contact_area, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE) 
        for i in range(len(contours)):
            # 最も大きい輪郭に外接矩形を作成し、画像に重ね書き
            ((xx,yy),(w, h),th) = cv2.minAreaRect(contours[i])

            if w*h > 1000000 : # >1000000 は大きい輪郭のみを残すため(330*330 ぐらい)
                hull = cv2.convexHull(contours[i]) #  凸包
                con_a_hull = cv2.cvtColor(contact_area, cv2.COLOR_GRAY2BGR)
                #img_co = cv2.drawContours(con_a_hull,[hull],0,(0,0,255),3) # 凸包
                #res_co = "/Users/BRLAB/others/img6/zz" + str(ikiti) +"_"+ str(x) +"co.jpg"
                #cv2.imwrite(res_co, img_co)

        # 凸包から四隅の座標を取得
        ue_y = 1500
        shita_y = 1500
        migi_x = 500
        hidari_x = 500
        for a in range(len(hull)): # まず上下左右を取得
            [p1] =  hull[a]
            if p1[1] < ue_y :
                ue_y = p1[1]
            if p1[0] > migi_x :
                migi_x = p1[0]
            if p1[1] > shita_y :
                shita_y = p1[1]
            if p1[0] < hidari_x :
                hidari_x = p1[0]
        kado = [[0 for i in range(2)] for j in range(4)] # 四隅[x,y]
        kado[0][:] = [hidari_x,ue_y]     # 左上
        kado[1][:] = [migi_x,ue_y]       # 右上
        kado[2][:] = [migi_x,shita_y]    # 右下
        kado[3][:] = [hidari_x, shita_y] # 左下

        '''
        # 四隅取得の確認
        #res_co = "/Users/BRLAB/others/img6/zz" + str(ikiti) +"_"+ str(x) +"co.jpg"
        #img_co = cv2.imread(res_co)
        point = 0
        while point < 4:
            point_0 = int(kado[point][0])
            point_1 = int(kado[point][1])
            print('p0 = ',point_0, 'p1 = ',point_1)
            cv2.putText(img_co,str(point), (point_0,point_1), cv2.FONT_HERSHEY_PLAIN,8, (0,0,255),3, cv2.LINE_8)
            point += 1
        cv2.imwrite(res_co, img_co)
        '''

        # RGB,edgeを取得する枠(200*200)を25 づつずらしていくので、その最終座標(左上)を計算
        amari_x = (migi_x-hidari_x) % 25
        last_p_x = migi_x - amari_x - 200
        amari_y = (shita_y-ue_y) % 25
        last_p_y = shita_y - amari_y - 200 
        
        # 枠の数（最大値）分、記録しておく行列作成
        frame_x = (last_p_x-hidari_x)/25
        frame_y = (last_p_y-ue_y)/25
        frame = int(frame_x*frame_y)
        xybgre_list = [[0 for i in range(7)] for j in range(frame)] # x,y,b,g,r,e,1e  xybgre_list[枠の番号(昇順)][x,y,b,g,r,e]
        
        # 初期値をゼロにする
        waku = 0
        for i in range(frame):
            kazu = 0
            for j in range(7):
                xybgre_list[waku][kazu] = 0
                kazu += 1
            waku += 1
    
        # point_one （枠の左上の座標）の初期値
        point_one = kado[0][:]
        img_point = contact_area[point_one[1]:point_one[1]+200, point_one[0]:point_one[0]+200]
        
        # いざ数値を取得！
        num_w = 0 # 枠の番号(昇順)
        while point_one[1] <= last_p_y :    
            while point_one[0] <= last_p_x :
                img_point = contact_area[point_one[1]:point_one[1]+200, point_one[0]:point_one[0]+200]
                atai = int(cv2.countNonZero(img_point))
                if atai == 40000:
                    # 座標記録
                    x_suti = int(point_one[0] + 50 )
                    y_suti = int(point_one[1] + 50 )
                    xybgre_list[num_w][0] = x_suti
                    xybgre_list[num_w][1] = y_suti
                    # BGR 取得___________________________________________________________________________     
                    yoso = 19*19
                    bgr_list = [0 for j in range(3)] # b,g,r
                    #bgr_list に BGR 取得, 0 入りの list に書き込む
                    ta = point_one[1] + 10 # 縦横、point の枠の10~190 の 19 点
                    ai = 0
                    while ta < point_one[1]+200 : 
                        ha = point_one[0] + 10
                        while ha < point_one[0]+200 :
                            [bl,gr,re] = sole_one[ta,ha,:]
                            bgr_list[0] += bl
                            bgr_list[1] += gr
                            bgr_list[2] += re
                            ai += 1
                            ha += 10
                        ta += 10
                    xybgre_list[num_w][2] = float(bgr_list[0]/yoso)
                    xybgre_list[num_w][3] = float(bgr_list[1]/yoso)
                    xybgre_list[num_w][4] = float(bgr_list[2]/yoso)
                    # edeg 取得___________________________________________________________________________     
                    edge_point = dst[point_one[1]:point_one[1]+200, point_one[0]:point_one[0]+200]
                    bunbo = edge_point.size
                    shiwa = cv2.countNonZero(edge_point)/bunbo
                    xybgre_list[num_w][5] = float(shiwa)
                    xybgre_list[num_w][6] = float(1/shiwa)

                    #print('x=',xybgre_list[num_w][0],' y=',xybgre_list[num_w][1],' b=', xybgre_list[num_w][2],
                    #    ' g=', xybgre_list[num_w][3],' r=', xybgre_list[num_w][4],' e=', xybgre_list[num_w][5])
                    num_w += 1
                point_one[0] += 25

            point_one[1] += 25 
            point_one[0] = kado[0][0]

        # ゼロの行列は消す
        waku = 0
        for i in range(frame):
            a = xybgre_list[waku][0]
            if a == 0:
                xybgre_list = np.delete(xybgre_list, waku, 0)
                #print('入った')
                waku -= 1
            waku += 1
        #print(xybgre_list)

        if x == 1:
            xybgre_list_1 = xybgre_list
            num_1 = num_w
        elif x == 2:
            xybgre_list_2 = xybgre_list
            num_2 = num_w
        # ________________________________________________________________________________________________________
    #繋ぐ___________________________________________________________

    i = 65161
    x = 1
    wr1 = "/Users/BRLAB/others/img6/dst" + str(name) +"_"+ str(x) +".jpg"
    i01 = cv2.imread(wr1)
    os.remove(wr1)
    wr_im = "/Users/BRLAB/others/img6/zz" + str(ikiti) + str(x) + ".jpg" # sole & mask
    os.remove(wr_im)

    x = 2
    wr2 = "/Users/BRLAB/others/img6/dst" + str(name) +"_"+ str(x) +".jpg"
    i02 = cv2.imread(wr2)
    os.remove(wr2)
    wr_im = "/Users/BRLAB/others/img6/zz" + str(ikiti) + str(x) + ".jpg"
    os.remove(wr_im)

    im_h = cv2.hconcat([i01, i02])
    res = "/Users/BRLAB/others/img6/dst" + str(name) +".jpg"
    cv2.imwrite(res, im_h)

    return xybgre_list_1,xybgre_list_2, num_1, num_2 # 左足と右足の　xybgre_list[枠の番号(昇順)][x,y,b,g,r,e], num_w[枠の番号(昇順],


def kaiki(im, pho, height, weight): # 取得した変数を基に、硬度分布と水分量分布マップを作る

    # height : 身長 (cm)
    # weight : 体重 (kg)
    # 接地面取得
    ikiti = 161
    area_con, name = area(im,pho, ikiti) # 接地面作成
    #suuzi = format(suuzi,'.1f')
    #print("接地面：",suuzi, "%")
    BMI = weight/height/height
    xybgre_list_1,  xybgre_list_2, num_1, num_2 = edge_rgb(name, ikiti) # 片足の「しわ」と「RGB」取得

    x = 0
    while x <= 2:
        x += 1
        if x == 1:
            xybgre_list = xybgre_list_1
            num_f = num_1
        elif x == 2:
            xybgre_list = xybgre_list_2
            num_f = num_2
        elif x == 3:
            break

        i = 65
        res_L = "/Users/BRLAB/others/img6/Lab_L" + str(name) +"_"+ str(x) +".jpg"
        area_hard = cv2.imread(res_L)
        area_wet = cv2.imread(res_L)

        xyHW_list0 = [[0 for i in range(4)] for j in range(num_f)] # xyHW_list[枠の番号(昇順)][x,y,Hard,Wet]
        HW_list0_0 = np.zeros((5,2), dtype=float)
        waku = 0
        a_x = 0
        a_y = 0
        if x == 1:
            koudo_max = 0
            koudo_min = 100
            suibun_max = 0
            suibun_min = 100
        for j in range(num_f):
            #200*200 の枠から得たデータを中央の50*50 の領域のデータとし、さらに50*50 の領域を25*25の大きさで４分割したものたちに与える。
            # num_f は 0 から num_f-1 までカウントされるのでこれでオッケー
            a_x = int(xybgre_list[waku][0])
            #print('a_x = ',a_x)
            xyHW_list0[waku][0] = a_x
            a_y = int(xybgre_list[waku][1])
            xyHW_list0[waku][1] = a_y

            # 硬度(Hardness) 推定式 [kakato_h] M_a_gr_1e を採用
            #切片	29.33359661
            #体重(kg)	-0.135720186
            #接地面new	0.054662043
            #かかとG	0.077699405
            #かかとR	-0.015930755
            #1/かかとしわ	-0.108131128
            hard =  29.33359661 -0.135720186*weight +0.054662043*area_con +0.077699405*xybgre_list[waku][3] -0.015930755*xybgre_list[waku][4] -0.108131128*xybgre_list[waku][6]
            xyHW_list0[waku][2] = hard
            if hard > koudo_max:
                koudo_max = hard
            if hard < koudo_min :
                koudo_min = hard
            if hard < 11:
                HW_list0_0[0][0] += 1
            if 11 <= hard < 21:
                HW_list0_0[1][0] += 1
            if 21 <= hard < 31:
                HW_list0_0[2][0] += 1
            if 31 <= hard < 41:
                HW_list0_0[3][0] += 1
            if 41 <= hard:
                HW_list0_0[4][0] += 1
            #print(hard)
            # 水分量(Wetness, 0~99) 推定式 [kakato_w] M_a_gr_1e を採用, 
            #切片	-15.11922676
            #体重(kg)	-0.167603347
            #接地面new	0.620639113
            #かかとG	-0.225582716
            #かかとR	0.206278907
            #1/かかとしわ	-0.105086363
            wet =  -15.11922676 -0.167603347*weight +0.620639113*area_con -0.225582716*xybgre_list[waku][3] +0.206278907*xybgre_list[waku][4] -0.105086363*xybgre_list[waku][6]        
            xyHW_list0[waku][3] = wet
            if wet > suibun_max:
                suibun_max = wet
            if wet < suibun_min:
                suibun_min = wet
            if wet < 11:
                HW_list0_0[0][1] += 1
            if 11 <= wet < 21:
                HW_list0_0[1][1] += 1
            if 21 <= wet < 31:
                HW_list0_0[2][1] += 1
            if 31 <= wet < 41:
                HW_list0_0[3][1] += 1
            if 41 <= wet:
                HW_list0_0[4][1] += 1
            # 硬度を 0~60, 水分量を 0~99 とする。
            hard = int(hard) # 硬度
            # 範囲は、初期値をh0, 差をhn とする
            h0 = 8
            hn = 4
            # 塗りつぶし
            #cv2.rectangle(img, pt1=(50, 150), pt2=(125, 250), color=(0, 255, 0), thickness=-1, lineType=cv2.LINE_4, shift=0)
            #cv2.rectangle(画像, 左上, 右下, 色, 線の厚みだがマイナス１で内側塗りつぶし, 線のタイプ, 座標における小数点以下を示すビット数)
            if hard < h0 :
                # iro = 1
                cv2.rectangle(area_hard, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(204, 224, 255), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif h0 <= hard < h0+1*hn  :
                # iro = 2
                cv2.rectangle(area_hard, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(153, 193, 255), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif h0+1*hn <= hard < h0+2*hn :
                # iro = 3
                cv2.rectangle(area_hard, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(108, 164, 248), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif h0+2*hn <= hard < h0+3*hn :
                # iro = 4
                cv2.rectangle(area_hard, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(67, 135, 238), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif h0+3*hn <= hard < h0+4*hn :
                # iro = 5
                cv2.rectangle(area_hard, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(30, 108, 224), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif h0+4*hn <= hard < h0+5*hn :
                # iro = 6
                cv2.rectangle(area_hard, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(32, 88, 171), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif h0+5*hn <= hard < h0+6*hn :
                # iro = 7
                cv2.rectangle(area_hard, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(30, 67, 122), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif h0+6*hn <= hard < h0+7*hn :
                # iro = 8
                cv2.rectangle(area_hard, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(24, 45, 77), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif h0+7*hn <= hard < h0+8*hn :
                # iro = 9
                cv2.rectangle(area_hard, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(14, 23, 36), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif h0+8*hn <= hard :
                # iro = 10
                cv2.rectangle(area_hard, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(0, 0, 0), thickness=-1, lineType=cv2.LINE_4, shift=0)
            
            wet = int(wet) # 水分量
            # 範囲は、初期値をw0, 差をwn とする
            w0 = 5
            wn = 5
            if wet < w0 :
                # iro = 10
                cv2.rectangle(area_wet, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(0, 0, 0), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif w0 <= wet < w0+1*wn :
                # iro = 9
                cv2.rectangle(area_wet, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(30, 32, 18), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif w0+1*wn <= wet < w0+2*wn :
                # iro = 8
                cv2.rectangle(area_wet, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(65, 73, 28), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif w0+2*wn <= wet < w0+3*wn :
                # iro = 7
                cv2.rectangle(area_wet, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(103, 116, 36), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif w0+3*wn <= wet < w0+4*wn :
                # iro = 6
                cv2.rectangle(area_wet, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(142, 163, 40), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif w0+4*wn <= wet < w0+5*wn :
                # iro = 5
                cv2.rectangle(area_wet, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(185, 214, 40), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif w0+5*wn <= wet < w0+6*wn :
                # iro = 4
                cv2.rectangle(area_wet, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(204, 230, 75), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif w0+6*wn <= wet < w0+7*wn :
                # iro = 3
                cv2.rectangle(area_wet, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(221, 242, 114), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif w0+7*wn <= wet < w0+8*wn :
                # iro = 2
                cv2.rectangle(area_wet, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(235, 250, 157), thickness=-1, lineType=cv2.LINE_4, shift=0)
            elif w0+8*wn <= wet :
                # iro = 1
                cv2.rectangle(area_wet, pt1=(a_x,a_y), pt2=(a_x+100,a_y+100), color=(246, 253, 204), thickness=-1, lineType=cv2.LINE_4, shift=0)

            waku += 1      

        if x == 1:
            HW_list0_1 = HW_list0_0
        if x == 1:
            HW_list0_2 = HW_list0_0
        hard_ad = "/Users/BRLAB/others/img6/hard"+ str(name)+ str(x) + ".jpg"
        cv2.imwrite(hard_ad, area_hard)
        wet_ad = "/Users/BRLAB/others/img6/wet"+ str(name)+ str(x) + ".jpg"
        cv2.imwrite(wet_ad, area_wet) 

    for q in range(5):
        for r in range(2):
            HW_1 = HW_list0_1[q][r]
            HW_2 = HW_list0_2[q][r]
            HW_list0_0[q][r] = ((HW_1 + HW_2) / (num_1 + num_2))*100
            HW_list0_0[q][r] = format(HW_list0_0[q][r],'.1f') # ratioの桁数指定,小数点第1位まで

    #print(HW_list0_0)
    x = 1
    wr1_h = "/Users/BRLAB/others/img6/hard"+ str(name)+ str(x) + ".jpg"
    i01_h = cv2.imread(wr1_h)
    os.remove(wr1_h)
    wr1_w = "/Users/BRLAB/others/img6/wet"+ str(name)+ str(x) + ".jpg"
    i01_w = cv2.imread(wr1_w)
    os.remove(wr1_w)
    wr_s = "/Users/BRLAB/others/img6/sole" + str(x) + ".jpg"
    sole_1 = cv2.imread(wr_s)
    os.remove(wr_s)
    res_L = "/Users/BRLAB/others/img6/Lab_L" + str(name) +"_"+ str(x) +".jpg"
    os.remove(res_L)
    wr_me = "/Users/BRLAB/others/img6/z_mask_e" + str(x) + ".jpg"
    os.remove(wr_me)

    x = 2
    wr2_h = "/Users/BRLAB/others/img6/hard"+ str(name)+ str(x) + ".jpg"
    i02_h = cv2.imread(wr2_h)
    wr2_w = "/Users/BRLAB/others/img6/wet"+ str(name)+ str(x) + ".jpg"
    i02_w = cv2.imread(wr2_w)

    # 横のバーと数字の書き込み++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # 硬度分布に書き込み
    sumple_x1= 960
    sumple_x2 = 1060
    iro = 0
    cv2.rectangle(i02_h, pt1=(840,20), pt2=(1160,620), color=(255, 255, 255), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 1
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_h, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(204, 224, 255), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 2
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_h, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(153, 193, 255), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 3
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_h, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(108, 164, 248), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 4
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_h, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(67, 135, 238), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 5
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_h, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(30, 108, 224), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 6
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_h, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(32, 88, 171), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 7
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_h, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(30, 67, 122), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 8
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_h, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(24, 45, 77), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 9
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_h, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(14, 23, 36), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 10
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_h, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(0, 0, 0), thickness=-1, lineType=cv2.LINE_4, shift=0)
    
    n = 0
    while n <= 8:
        text_h = h0 + n*hn 
        y = int( 165 + n*50 )
        if text_h < 10 :
            cv2.putText(i02_h,str(text_h), (1110, y), cv2.FONT_HERSHEY_PLAIN,3, (0,0,0),3, cv2.LINE_8)
        else :
            cv2.putText(i02_h,str(text_h), (1080, y), cv2.FONT_HERSHEY_PLAIN,3, (0,0,0),3, cv2.LINE_8)
        n += 1
    text_h1 = 'Hardness'
    cv2.putText(i02_h,str(text_h1), (855, 80), cv2.FONT_HERSHEY_PLAIN,4, (0,0,0),5, cv2.LINE_8)
    # 縦
    cv2.line(i02_h, pt1=(925, 97), pt2=(925, 593), color=(0,0,0), thickness=3, lineType=cv2.LINE_8, shift=0)
    # 横
    cv2.line(i02_h, pt1=(905, 97), pt2=(945, 97), color=(0,0,0), thickness=3, lineType=cv2.LINE_8, shift=0)
    cv2.line(i02_h, pt1=(910, 195), pt2=(940, 195), color=(0,0,0), thickness=3, lineType=cv2.LINE_8, shift=0)
    cv2.line(i02_h, pt1=(910, 445), pt2=(940, 445), color=(0,0,0), thickness=3, lineType=cv2.LINE_8, shift=0)
    cv2.line(i02_h, pt1=(905, 593), pt2=(945, 593), color=(0,0,0), thickness=3, lineType=cv2.LINE_8, shift=0)
    # text
    cv2.putText(i02_h,'S', (860, 170), cv2.FONT_HERSHEY_PLAIN,4, (0,0,0),5, cv2.LINE_8)
    cv2.putText(i02_h,'N', (860, 340), cv2.FONT_HERSHEY_PLAIN,4, (0,0,0),5, cv2.LINE_8)
    cv2.putText(i02_h,'H', (860, 540), cv2.FONT_HERSHEY_PLAIN,4, (0,0,0),5, cv2.LINE_8)

    #############################################################################
    # 水分量分布に書き込み
    sumple_x1= 960
    sumple_x2 = 1060
    iro = 0
    cv2.rectangle(i02_w, pt1=(840,20), pt2=(1160,620), color=(255, 255, 255), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 1
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_w, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(246, 255, 204), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 2
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_w, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(235, 250, 157), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 3
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_w, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(221, 242, 114), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 4
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_w, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(204, 230, 75), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 5
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_w, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(185, 214, 40), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 6
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_w, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(142, 163, 40), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 7
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_w, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(103, 116, 36), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 8
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_w, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(65, 73, 28), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 9
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_w, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(30, 32, 18), thickness=-1, lineType=cv2.LINE_4, shift=0)
    iro = 10
    sumple_y1 = int(95 + (iro-1)*50)
    sumple_y2 = int(sumple_y1 + 50)
    cv2.rectangle(i02_w, pt1=(sumple_x1, sumple_y1), pt2=(sumple_x2, sumple_y2), color=(0, 0, 0), thickness=-1, lineType=cv2.LINE_4, shift=0)

    n = 0
    while n <= 8:
        text_w = int(w0 + (8-n)*wn) 
        y = int( 165 + n*50 )
        if text_w < 10 :
            cv2.putText(i02_w,str(text_w), (1110, y), cv2.FONT_HERSHEY_PLAIN,3, (0,0,0),3, cv2.LINE_8)
        else :
            cv2.putText(i02_w,str(text_w), (1080, y), cv2.FONT_HERSHEY_PLAIN,3, (0,0,0),3, cv2.LINE_8)
        n += 1
    text_w1 = 'Wetness'
    cv2.putText(i02_w,str(text_w1), (865, 80), cv2.FONT_HERSHEY_PLAIN,4, (0,0,0),5, cv2.LINE_8)
    # 縦
    cv2.line(i02_w, pt1=(925, 97), pt2=(925, 593), color=(0,0,0), thickness=3, lineType=cv2.LINE_8, shift=0)
    # 横
    cv2.line(i02_w, pt1=(905, 97), pt2=(945, 97), color=(0,0,0), thickness=3, lineType=cv2.LINE_8, shift=0)
    cv2.line(i02_w, pt1=(910, 195), pt2=(940, 195), color=(0,0,0), thickness=3, lineType=cv2.LINE_8, shift=0)
    cv2.line(i02_w, pt1=(910, 445), pt2=(940, 445), color=(0,0,0), thickness=3, lineType=cv2.LINE_8, shift=0)
    cv2.line(i02_w, pt1=(905, 593), pt2=(945, 593), color=(0,0,0), thickness=3, lineType=cv2.LINE_8, shift=0)
    # text
    cv2.putText(i02_w,'M', (860, 170), cv2.FONT_HERSHEY_PLAIN,4, (0,0,0),5, cv2.LINE_8)
    cv2.putText(i02_w,'N', (860, 340), cv2.FONT_HERSHEY_PLAIN,4, (0,0,0),5, cv2.LINE_8)
    cv2.putText(i02_w,'D', (860, 540), cv2.FONT_HERSHEY_PLAIN,4, (0,0,0),5, cv2.LINE_8)

    # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    os.remove(wr2_h)
    os.remove(wr2_w)
    wr_s = "/Users/BRLAB/others/img6/sole" + str(x) + ".jpg"
    sole_2 = cv2.imread(wr_s)
    os.remove(wr_s)
    res_L = "/Users/BRLAB/others/img6/Lab_L" + str(name) +"_"+ str(x) +".jpg"
    os.remove(res_L)
    wr_me = "/Users/BRLAB/others/img6/z_mask_e" + str(x) + ".jpg"
    os.remove(wr_me)

    im_h = cv2.hconcat([i01_h, i02_h])
    res_h = "/Users/BRLAB/others/img6/result_6hard_"+ str(name)+".jpg"
    cv2.imwrite(res_h, im_h)
    im_w = cv2.hconcat([i01_w, i02_w])
    res_w = "/Users/BRLAB/others/img6/result_6wet_"+ str(name)+".jpg"
    cv2.imwrite(res_w, im_w)   
    im_s = cv2.hconcat([sole_1, sole_2])
    wr_s = "/Users/BRLAB/others/img6/sole_" + str(name) + ".jpg"
    cv2.imwrite(wr_s, im_s)
    '''
    im_h = cv2.cvtColor(im_h, cv2.COLOR_BGR2RGB)
    plt.imshow(im_h)
    plt.show()
    im_w = cv2.cvtColor(im_w, cv2.COLOR_BGR2RGB)
    plt.imshow(im_w)
    plt.show()
    '''

    return koudo_max, koudo_min, suibun_max, suibun_min, HW_list0_0,im_h, im_w



def all(im, pho, hi, we):
    numbe = pho
    numbe = numbe[:3] # 先頭から二文字抽出
    numbe2 = numbe[1:2] # 二文字目
    numbe3 = numbe[2:3] # 三文字目
    if numbe2 == '_':
        numbe_1 = numbe[:1] # 一文字目まで
        numbe = numbe_1
    elif numbe3 == '_':
        numbe_2 = numbe[:2] # 二文字目まで
        numbe = numbe_2
    else :
        numbe = numbe
    
    e_num = numbe
    number = int(numbe) + 1 # No. + 1
    '''
    # excel ファイルの読み込み
    book_date = px.load_workbook('/Users/BRLAB/Desktop/photo_20220204/hikensya_20220204made.xlsx') 
    sheet_date1 = book_date['Sheet1']
    NaN = 'NaN'
    years = sheet_date1.cell(row=number,column=6).value
    hi = sheet_date1.cell(row=number,column=7).value # 身長、 [D29」セルは、「行29、列4」
    we = sheet_date1.cell(row=number,column=8).value # 体重、[D29」セルは、「行29、列4」
    '''
    kou_max, kou_min, sui_max, sui_min, HW_list,im_h, im_w = kaiki(im, pho, hi, we)

    wb = px.load_workbook('/Users/BRLAB/others/img6/Book1.xlsx')
    ws = wb.active
    # セルに書き込み
    ws.cell(number, 1).value = int(e_num)
    '''
    ws.cell(number, 2).value = int(years)
    '''
    ws.cell(number, 3).value = int(hi)
    ws.cell(number, 4).value = int(we)

    ws.cell(number, 5).value = int(kou_max)
    ws.cell(number, 6).value = int(kou_min)
    ws.cell(number, 7).value = int(sui_max)
    ws.cell(number, 8).value = int(sui_min)
    ws.cell(number, 10).value = float(HW_list[0][0])
    ws.cell(number, 11).value = float(HW_list[1][0])
    ws.cell(number, 12).value = float(HW_list[2][0])
    ws.cell(number, 13).value = float(HW_list[3][0])
    ws.cell(number, 14).value = float(HW_list[4][0])
    ws.cell(number, 16).value = float(HW_list[0][1])
    ws.cell(number, 17).value = float(HW_list[1][1])
    ws.cell(number, 18).value = float(HW_list[2][1])
    ws.cell(number, 19).value = float(HW_list[3][1])
    ws.cell(number, 20).value = float(HW_list[4][1])

    wb.save('/Users/BRLAB/others/img6/Book1.xlsx')

    return HW_list, im_h, im_w

    
'''
pho = '1_1_7520514_20211118.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
#im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
hi = 170
we = 60
HW_list, im_h, im_w = all(im, pho, hi, we)
print(HW_list)

pho = '5_2_7521535_20211118.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '6_2_7521504_20211118.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '7_2_7520504_20211118.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '8_2_7521506_20211118.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '6_2_7521504_20211118.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '9_2_7520509_20211118.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '10_2_7500001_20211118.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '22_2_7520514_20211119.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '23_2_7521506_20211119.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '24_2_7518084_20211119.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '25_2_7520509_20211119.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '26_2_7520001_20211119.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '27_2_7518043_20211119.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180) # 上下回転
all(im, pho)

pho = '54_2_7518084_20211126.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '116_1_7520002_20211215.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '117_1_7520003_20211216.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '118_1_7520004_20211216.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '119_1_7520005_20211216.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '120_1_7520006_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '121_1_2601_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '122_1_2602_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '123_1_2603_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '124_1_2604_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '125_1_2605_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '126_1_2606_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '127_1_2607_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '128_1_2608_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '129_1_2609_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '130_1_2610_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '131_1_2611_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '132_1_2612_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '135_1_2615_20211220.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '136_1_2616_20211220.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '137_1_2617_20211220.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '138_1_2618_20211220.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '139_1_2619_20211220.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)

pho = '133_1_2613_20211217.jpg'
date = '/Users/BRLAB/Desktop/photo_20220204/' + str(pho)
im = cv2.imread(date) 
im = cv2.rotate(im, cv2.ROTATE_180)
all(im, pho)
'''
