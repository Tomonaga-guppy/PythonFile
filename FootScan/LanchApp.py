# いろんなモジュールを import
# 逆にここにあるモジュールを cmd で pip install すればPC変えてもこのプログラムをつかって装置動かせる。あと保存先のファイルとEPSONアプリ必要だけど、、
import datetime
import tkinter as tk #GUIアプリケーションで使う
import tkinter.ttk as ttk
from tkinter import filedialog as fd  # 参照用
import cv2  # 画像解析
import openpyxl as px  # excel 書き込み用、pip install openpyxl 必要

#解析用自作ファイル
import kaiseki_press
import kaiseki_area
import kaiseki_jushin_angl
import kaiseki_edge
import kaiseki_HandW

import win32api
import numpy as np # pip install numpy 必
from matplotlib import pyplot as plt # pip install matplotlib 必要
import serial,time # serialは、PythonとArduinoなどの機器との間でシリアル通信を実現させられるモジュール、pip install pyserial 必要
# シリアル通信は接続されたコンピュータ同士で、1ビットずつデータを送る通信の方式,パラレル通信はデータを複数のビットにまとめて送る通信の方式
from multiprocessing import Process # マルチプロセスによる並列処理をサポートする
import threading # マルチスレッドプログラミングを行うためのモジュール
import csv # CSVファイル(Comma Separated Valueの略で、カンマ（,）で区切られた値が含まれているテキストファイル)の読み書きをするためのモジュール
import glob # 引数に指定されたパターンにマッチするファイルパス名を取得、処理するときに使うモジュール
import os # Pythonのコード上でOS（オペレーティング・システム）に関する操作を実現するためのモジュール,pip install pywin32 必要？
from scipy import signal
import pandas as pd
from PIL import Image, ImageTk ,ImageWin # Pillow の pip install が必要
import win32print # 印刷に必要、pip install pywin32 で入る。~~~~~~~~ のままで大丈夫。

wb = px.load_workbook('/Users/BRLAB/hikensya.xlsx') # 被験者情報の保存先のエクセルファイル
ws = wb.active # ファイルの書き換え可能にするactive
ser = serial.Serial('COM3',9600) #COM4(Arduino)とシリアル通信を行う。
time.sleep(3) # 3秒スリープ
ser.write(b'2') # シリアル通信でArduinoにcase2の指令
time.sleep(40) # 40秒スリープ

# ここから被験者番号入力 #####################################################################
def id(): # 被検者番号入力用GUIを一つの関数に
  global frame_id # frame_id(被験者番号の入力値) を他と共有できるように、、global!
  frame_id=0
  # 被検者番号の数字入力用ボタンを押したあとの動作を定義、click_btn1…はあとで使う
  def click_btn1(): # 159行目のbutton1("1"のボタン)を押されたときのイベント
      if label['text'] == '0': # label は156行目で定義した文字表示用ラベル、これを「1」に、(47行目)もしくは「元の数字+1」にする
          label['text'] = '1'
          buttonEnter['fg'] = "red" # buttonEnterは192行目で定義した「決定」のボタン、赤色に染めるエフェクト?
          label2['fg']="black" # label2は153行目に定義した「被検者番号を入力してください」のラベル
      else:
          label['text'] = label['text'] +'1'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn2():
      if label['text'] == '0':
          label['text'] = '2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn3():
      if label['text'] == '0':
          label['text'] = '3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn4():
      if label['text'] == '0':
          label['text'] = '4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn5():
      if label['text'] == '0':
          label['text'] = '5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn6():
      if label['text'] == '0':
          label['text'] = '6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn7():
      if label['text'] == '0':
          label['text'] = '7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn8():
      if label['text'] == '0':
          label['text'] = '8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn9():
      if label['text'] == '0':
          label['text'] = '9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn0():
      if label['text'] == '0':
          label['text'] = '0'
      else:
          label['text'] = label['text'] +'0'
          buttonEnter['fg'] = "red"
          label2['fg']="black"

  def click_btnAc():
      label['text'] = '0'
      buttonEnter['fg'] = "black"
      label2['fg']="black"

  def click_btnFinish():
    frame_id.destroy()
    start_jouken()

  def click_btnEnter():
    global idnum
    idnum=0
    idnum=int(label['text'])
    print(idnum)
    frame_id.destroy()
    ori()

  baseGround.title('足裏測定')

  # フレームサイズ定義
  frame_id = tk.Frame(baseGround, width=500, height=600)
  frame_id.pack() # 配置をpackに
  # ラベルやボタンの作成と配置
  label2 = tk.Label(frame_id, text="被検者番号を入力してください",font=("",30))
  label2.place(x=25, y=0)

  label = tk.Label(frame_id, text='0', background='lightgray',font=("",50))
  label.place(x=50, y=50, width=400, height=105)

  button1 = tk.Button(frame_id,text = '1', command=click_btn1,font=("",50)) # 上で定義したclick_btn1…を装備
  button1.place(x=50, y=205, relwidth=0.18, relheight= 0.125)

  button2 = tk.Button(frame_id,text = '2', command=click_btn2,font=("",50)) # 上で定義したclick_btn2…を装備
  button2.place(x=153, y=205, relwidth=0.18, relheight= 0.125)

  button3 = tk.Button(frame_id,text = '3', command=click_btn3,font=("",50)) #
  button3.place(x=256, y=205, relwidth=0.18, relheight= 0.125)

  button4 = tk.Button(frame_id,text = '4', command=click_btn4,font=("",50))
  button4.place(x=50, y=295, relwidth=0.18, relheight= 0.125)

  button5 = tk.Button(frame_id,text = '5', command=click_btn5,font=("",50))
  button5.place(x=153, y=295, relwidth=0.18, relheight= 0.125)

  button6 = tk.Button(frame_id,text = '6', command=click_btn6,font=("",50))
  button6.place(x=256, y=295, relwidth=0.18, relheight= 0.125)

  buttonFinish = tk.Button(frame_id,text = '終了', command=click_btnFinish,font=("",15))
  buttonFinish.place(x=359, y=295, relwidth=0.18, relheight= 0.125)

  button7 = tk.Button(frame_id,text = '7', command=click_btn7,font=("",50))
  button7.place(x=50, y=385, relwidth=0.18, relheight= 0.125)

  button8 = tk.Button(frame_id,text = '8', command=click_btn8,font=("",50))
  button8.place(x=153, y=385, relwidth=0.18, relheight= 0.125)

  button9 = tk.Button(frame_id,text = '9', command=click_btn9,font=("",50))
  button9.place(x=256, y=385, relwidth=0.18, relheight= 0.125)

  button0 = tk.Button(frame_id,text = '0', command=click_btn0,font=("",50))
  button0.place(x=153, y=475, relwidth=0.18, relheight= 0.125)

  buttonEnter = tk.Button(frame_id,text = '決定', command=click_btnEnter,font=("",50))
  buttonEnter.place(x=256, y=475, relwidth=0.38, relheight= 0.125)

  buttonAc = tk.Button(frame_id,text = 'AC', command=click_btnAc,font=("",50))
  buttonAc.place(x=50, y=475, relwidth=0.18, relheight= 0.125)

# ここから年齢入力 #########################################################################
def make():
  global frame
  # 年齢の数字入力用ボタンを押したあとの動作を定義、click_btn1…はあとで使う
  def click_btn1():
      if label['text'] == '0':
          label['text'] = '1'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'1'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn2():
      if label['text'] == '0':
          label['text'] = '2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn3():
      if label['text'] == '0':
          label['text'] = '3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn4():
      if label['text'] == '0':
          label['text'] = '4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn5():
      if label['text'] == '0':
          label['text'] = '5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn6():
      if label['text'] == '0':
          label['text'] = '6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn7():
      if label['text'] == '0':
          label['text'] = '7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn8():
      if label['text'] == '0':
          label['text'] = '8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn9():
      if label['text'] == '0':
          label['text'] = '9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn0():
      if label['text'] == '0':
          label['text'] = '0'
      else:
          label['text'] = label['text'] +'0'
          buttonEnter['fg'] = "red"
          label2['fg']="black"

  def click_btnAc():
      label['text'] = '0'
      buttonEnter['fg'] = "black"
      label2['fg']="black"

  def click_btnReturn():
    frame.destroy()
    ori()

  def click_btnFinish():
    frame.destroy()
    start_jouken()

  def click_btnEnter():
    global nenrei
    nenrei=0

    if( int(label['text']) > 15 and int(label['text']) < 120 ):
        nenrei=int(label['text'])
        frame.destroy()
        change()
    else:
        label2['fg']="red"
        buttonEnter['fg'] = "black"


  baseGround.title('足裏測定')

  frame = tk.Frame(baseGround, width=500, height=600)
  frame.pack()
  # ラベルやボタンの作成と配置
  label2 = tk.Label(frame, text="年齢を入力してください",font=("",30))
  label2.place(x=40, y=0)

  label = tk.Label(frame, text='0', background='lightgray',font=("",50))
  label.place(x=50, y=50, width=400, height=105)

  button1 = tk.Button(frame,text = '1', command=click_btn1,font=("",50))
  button1.place(x=50, y=205, relwidth=0.18, relheight= 0.125)

  button2 = tk.Button(frame,text = '2', command=click_btn2,font=("",50))
  button2.place(x=153, y=205, relwidth=0.18, relheight= 0.125)

  button3 = tk.Button(frame,text = '3', command=click_btn3,font=("",50))
  button3.place(x=256, y=205, relwidth=0.18, relheight= 0.125)

  buttonReturn = tk.Button(frame,text = '性別入力\nに戻る', command=click_btnReturn,font=("",15))
  buttonReturn.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

  button4 = tk.Button(frame,text = '4', command=click_btn4,font=("",50))
  button4.place(x=50, y=295, relwidth=0.18, relheight= 0.125)

  button5 = tk.Button(frame,text = '5', command=click_btn5,font=("",50))
  button5.place(x=153, y=295, relwidth=0.18, relheight= 0.125)

  button6 = tk.Button(frame,text = '6', command=click_btn6,font=("",50))
  button6.place(x=256, y=295, relwidth=0.18, relheight= 0.125)

  buttonFinish = tk.Button(frame,text = '終了', command=click_btnFinish,font=("",15))
  buttonFinish.place(x=359, y=295, relwidth=0.18, relheight= 0.125)

  button7 = tk.Button(frame,text = '7', command=click_btn7,font=("",50))
  button7.place(x=50, y=385, relwidth=0.18, relheight= 0.125)

  button8 = tk.Button(frame,text = '8', command=click_btn8,font=("",50))
  button8.place(x=153, y=385, relwidth=0.18, relheight= 0.125)

  button9 = tk.Button(frame,text = '9', command=click_btn9,font=("",50))
  button9.place(x=256, y=385, relwidth=0.18, relheight= 0.125)

  button0 = tk.Button(frame,text = '0', command=click_btn0,font=("",50))
  button0.place(x=153, y=475, relwidth=0.18, relheight= 0.125)

  buttonEnter = tk.Button(frame,text = '決定', command=click_btnEnter,font=("",50))
  buttonEnter.place(x=256, y=475, relwidth=0.38, relheight= 0.125)

  buttonAc = tk.Button(frame,text = 'AC', command=click_btnAc,font=("",50))
  buttonAc.place(x=50, y=475, relwidth=0.18, relheight= 0.125)


# ここから身長入力 ######################################################################
def change():
  global frame_app
  global jushinlist
  # 身長の数字入力用ボタンを押したあとの動作を定義、click_btn1…はあとで使う
  def click_btn1():
      if label['text'] == '0':
          label['text'] = '1'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'1'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn2():
      if label['text'] == '0':
          label['text'] = '2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn3():
      if label['text'] == '0':
          label['text'] = '3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn4():
      if label['text'] == '0':
          label['text'] = '4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn5():
      if label['text'] == '0':
          label['text'] = '5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn6():
      if label['text'] == '0':
          label['text'] = '6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn7():
      if label['text'] == '0':
          label['text'] = '7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn8():
      if label['text'] == '0':
          label['text'] = '8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn9():
      if label['text'] == '0':
          label['text'] = '9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn0():
      if label['text'] == '0':
          label['text'] = '0'
      else:
          label['text'] = label['text'] +'0'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btnAc():
      label['text'] = '0'
      buttonEnter['fg'] = "black"
      label2['fg']="black"
  def click_btnReturn(): # 戻るボタンで年齢入力に戻る
    frame_app.destroy()
    make()
  def click_btnFinish(): # 終了ボタンでstart_jouken() 一番初めに戻る
    frame_app.destroy()
    start_jouken()
  def click_btnEnter(): # 決定ボタンで、被験者情報をエクセルに書き込み、体重測定開始
    global today
    global shincho
    global enumber
    enumber=1

    if( int(label['text']) > 100 and int(label['text']) < 250 ):
        shincho=int(label['text'])

        dt_now = datetime.datetime.now() # 今日の日付
        yearnum=dt_now.year
        monthnum=dt_now.month
        monthnum = '{0:0>2}'.format(monthnum)
        daynum=dt_now.day
        daynum = '{0:0>2}'.format(daynum)
        today=str(yearnum)+str(monthnum)+str(daynum)

        while not ws.cell(enumber, 1).value is None: # hikensya.xlsx でデータの無い行まで１からプラス１していく（必然と最後尾に付く）
            enumber += 1
        ws.cell(enumber, 1).value = int(enumber)-1 # ここから下、被験者情報をエクセルに保存、cell(enumber, 1)はenumber行 1 列目のセルを指定
        ws.cell(enumber, 2).value = int(today)
        ws.cell(enumber, 3).value = int(jnum)
        ws.cell(enumber, 4).value = int(idnum)
        ws.cell(enumber, 5).value = int(seibetsu)
        ws.cell(enumber, 6).value = int(nenrei)
        ws.cell(enumber, 7).value = int(shincho)
        #ws.cell(enumber, 8).value = float(weight)
        wb.save('hikensya.xlsx')


        frame_app.destroy()
        taijucali() # 体重測定に移行するようコール
        # start()
        # baseGround.destroy()
    else:
        label2['fg']="red"
        buttonEnter['fg'] = "black"

  baseGround.title('足裏測定')

  frame_app = tk.Frame(baseGround, width=500, height=600)
  frame_app.pack()
  # ラベルやボタンの作成と配置

  label2 = tk.Label(frame_app, text="身長(cm)を入力してください",font=("",30))
  label2.place(x=35, y=0)

  label = tk.Label(frame_app, text='0', background='lightgray',font=("",50))
  label.place(x=50, y=50, width=400, height=105)

  button1 = tk.Button(frame_app,text = '1', command=click_btn1,font=("",50))
  button1.place(x=50, y=205, relwidth=0.18, relheight= 0.125)

  button2 = tk.Button(frame_app,text = '2', command=click_btn2,font=("",50))
  button2.place(x=153, y=205, relwidth=0.18, relheight= 0.125)

  button3 = tk.Button(frame_app,text = '3', command=click_btn3,font=("",50))
  button3.place(x=256, y=205, relwidth=0.18, relheight= 0.125)

  buttonReturn = tk.Button(frame_app,text = '年齢入力\nに戻る', command=click_btnReturn,font=("",15))
  buttonReturn.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

  button4 = tk.Button(frame_app,text = '4', command=click_btn4,font=("",50))
  button4.place(x=50, y=295, relwidth=0.18, relheight= 0.125)

  button5 = tk.Button(frame_app,text = '5', command=click_btn5,font=("",50))
  button5.place(x=153, y=295, relwidth=0.18, relheight= 0.125)

  button6 = tk.Button(frame_app,text = '6', command=click_btn6,font=("",50))
  button6.place(x=256, y=295, relwidth=0.18, relheight= 0.125)

  buttonFinish = tk.Button(frame_app,text = '終了', command=click_btnFinish,font=("",15))
  buttonFinish.place(x=359, y=295, relwidth=0.18, relheight= 0.125)

  button7 = tk.Button(frame_app,text = '7', command=click_btn7,font=("",50))
  button7.place(x=50, y=385, relwidth=0.18, relheight= 0.125)

  button8 = tk.Button(frame_app,text = '8', command=click_btn8,font=("",50))
  button8.place(x=153, y=385, relwidth=0.18, relheight= 0.125)

  button9 = tk.Button(frame_app,text = '9', command=click_btn9,font=("",50))
  button9.place(x=256, y=385, relwidth=0.18, relheight= 0.125)

  button0 = tk.Button(frame_app,text = '0', command=click_btn0,font=("",50))
  button0.place(x=153, y=475, relwidth=0.18, relheight= 0.125)

  buttonEnter = tk.Button(frame_app,text = '決定', command=click_btnEnter,font=("",50))
  buttonEnter.place(x=256, y=475, relwidth=0.38, relheight= 0.125)

  buttonAc = tk.Button(frame_app,text = 'AC', command=click_btnAc,font=("",50))
  buttonAc.place(x=50, y=475, relwidth=0.18, relheight= 0.125)

# ここから性別入力 ##############################################################
def ori():
    def click_btn1():
        if label['text'] == '女':
            label['text'] = '男'
            buttonEnter['fg'] = "red"
            label2['fg']="black"

        else:
            label['text'] = '男'
            buttonEnter['fg'] = "red"
            label2['fg']="black"

    def click_btn2():
        if label['text'] == '男':
            label['text'] = '女'
            buttonEnter['fg'] = "red"
            label2['fg']="black"

        else:
            label['text'] = '女'
            buttonEnter['fg'] = "red"
            label2['fg']="black"

    def click_btnFinish():
        frame_ori.destroy()
        start_jouken()

    def click_btnEnter():
        global seibetsu
        seibetsu=0
        if label['text'] == '男':
            seibetsu=1
            frame_ori.destroy()
            make()
        elif label['text'] == '女':
            seibetsu=2
            frame_ori.destroy()
            make()
        else:
            label2['fg']="red"

    frame_ori = tk.Frame(baseGround, width=500, height=600)
    frame_ori.pack()
    # ラベルやボタンの作成と配置

    label2 = tk.Label(frame_ori, text="性別を入力してください",font=("",30))
    label2.place(x=40, y=0)

    label = tk.Label(frame_ori, text='', background='lightgray',font=("",50))
    label.place(x=50, y=50, width=400, height=105)

    button1 = tk.Button(frame_ori,text = '男', command=click_btn1,font=("",50))
    button1.place(x=50, y=205, relwidth=0.18, relheight= 0.125)

    button2 = tk.Button(frame_ori,text = '女', command=click_btn2,font=("",50))
    button2.place(x=153, y=205, relwidth=0.18, relheight= 0.125)

    buttonEnter = tk.Button(frame_ori,text = '決定', command=click_btnEnter,font=("",50))
    buttonEnter.place(x=256, y=205, relwidth=0.36, relheight= 0.125)

    buttonFinish = tk.Button(frame_ori,text = '終了', command=click_btnFinish)
    buttonFinish.place(x=359, y=295, relwidth=0.18, relheight= 0.125)


# 画像処理/解析 GUI #################################################################
def kaiseki_gui():
    # メインページを作成
    main_frame = tk.Frame(baseGround,width=1410, height=698)
    main_frame.place(x=1,y = 1)

    # テキストラベル１
    file_name = tk.StringVar() # 後に値を変えられるテキストラベル
    file_name.set("画像 ： 未選択です")
    label = tk.Label(main_frame,textvariable=file_name, font=('', 12))

    # テキストラベル２
    kaiseki_chu = tk.StringVar() # 後に値を変えられるテキストラベル
    kaiseki_chu.set("")
    label2 = tk.Label(main_frame,textvariable=kaiseki_chu, font=('', 11))

    # 解析用画像データ
    date = tk.StringVar()
    date.set("")

    def hazimenimodoru():
        main_frame.destroy()
        start_jouken()

    # 参照ボタン押下後のevent。//////////////////////////////////////////////////////////////
    def file_dialog():
        global file_name_ans
        # 参照
        fTyp = [("", "*")]
        iDir = os.path.abspath(os.path.dirname(__file__))
        file_name_ans = fd.askopenfilename(filetypes=fTyp, initialdir=iDir)
        if len(file_name_ans) == 0:
            file_name.set("ファイルを選択してください")
        else:
            file_name.set("画像 ： " + file_name_ans)
        # 参照で選択した画像表示（JPNG のみ）
        global img
        global canvas1
        img = Image.open(file_name_ans, 'r')
        w = img.width # 横幅を取得
        h = img.height # 縦幅を取得
        img = img.resize(( int(w * (550/h)), int(h * (550/h)) ))
        img = ImageTk.PhotoImage(img)
        canvas1 = tk.Canvas(main_frame,width=500, height=600)
        canvas1.place( x = 10, y = 90)
        canvas1.create_image(240, 275, image=img)

    #//////////////////////////////////////////////////////////////////////////////////////

    # ファイルボタン
    file_button = tk.Button(main_frame, text='ファイル', width=8, height=1, command=file_dialog)
    file_button.place( x = 3, y = 3 )
    canvas1 = tk.Canvas(main_frame,width=500, height=600)

    file_button_end = tk.Button(main_frame, text='はじめに戻る', width=10, height=1, command=hazimenimodoru)
    file_button_end.place( x = 80, y = 3)

    # "解析"ボタンを押したあとのイベント
    # 解析結果ウィンドウの生成 *********************************************************
    # kaiseki_openpyxl.py より

    def kaiseki():

        # 画像解析部分wwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwww
        global datekey # 画像処理する画像の住所
        global im # 画像処理する画像本体
        global imgkai
        global numbe
        global img1
        global img2, value7
        global img3, value6
        global img4, value8
        global file_name_ans
        global img01, zyusyo1
        global img02, zyusyo2
        global img03, zyusyo3
        global img04, zyusyo4
        global img05, zyusyo5
        global img06, zyusyo6

        print("処理中...")

        # 画像のファイル名にあるはじめの数字を取得したい
        datekey = file_name_ans.lstrip("C"":") # 画像の住所
        numbe_num = len(datekey) # 文字数
        nu=1
        while not datekey[numbe_num-nu] == '/':
            nu += 1
        numbe_nu = numbe_num - nu + 1
        numbe = datekey[numbe_nu:numbe_num]
        #print(numbe)
        pho = numbe.strip(".jpg")


        # numbe を画像名の最初の数字にしたい
        numbe = numbe[:3] # 先頭から二文字抽出
        numbe2 = numbe[1:2] # 二文字目
        numbe3 = numbe[2:3] # 三文字目
        if numbe2 == '_':
            numbe_1 = numbe[:1] # 一文字目まで
            numbe = numbe_1
        elif numbe3 == '_':
            numbe_2 = numbe[:2] # 二文字目まで
            numbe = numbe_2
        else :
            numbe = numbe

        # 解析１
        datekey = file_name_ans.lstrip("C"":") # 画像の住所
        im = cv2.imread(datekey) # 足裏画像
        img1 = kaiseki_jushin_angl.juangl(im,numbe,pho) # 足裏画像に重心動揺起動をプロット kaiseki_jushin_angl.py を使用

        kaiseki_chu.set("処理中.")

        # 解析２ (これは映えるから入れている)
        im = cv2.imread(datekey) # 足裏画像
        img2, value7 = kaiseki_press.press(im,numbe) # 足裏画像の模擬圧力分布画像を取得 kaiseki_press.py を使用
        value7 = format(value7,'.1f') # 数値の桁数指定

        kaiseki_chu.set("処理中..")

        # 解析３
        im = cv2.imread(datekey) # 足裏画像
        img3, value6 = kaiseki_area.area(im,numbe) # 足裏画像の接地面画像を取得 kaiseki_area.py を使用
        value6 = format(value6,'.1f') # 数値の桁数指定

        kaiseki_chu.set("処理中...")

        # 解析４
        im = cv2.imread(datekey) # 足裏画像
        img4, value8 = kaiseki_edge.edg(im,numbe) # 足裏画像のしわひび割れ画像を取得 kaiseki_edge.py を使用
        value8 = format(value8,'.2f') # 数値の桁数指定

        kaiseki_chu.set("処理中....")

        #wwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwww

        # excel ファイルの読み込み（エクセルファイルから値を取得）
        book_date = px.load_workbook('/Users/BRLAB/hikensya.xlsx')
        sheet_date1 = book_date['Sheet1']
        # 値を取得
        number = int(numbe) + 1 # No. + 1
        value1 = sheet_date1.cell(row=number,column=4).value # ID、[D29」セルは、「行29、列4」
        value2 = sheet_date1.cell(row=number,column=5).value # 性別、 [D29」セルは、「行29、列4」
        value2 = int(value2)
        value3 = sheet_date1.cell(row=number,column=6).value # 年齢、[D29」セルは、「行29、列4」
        value4 = sheet_date1.cell(row=number,column=7).value # 身長、 [D29」セルは、「行29、列4」
        value5 = sheet_date1.cell(row=number,column=8).value # 体重、[D29」セルは、「行29、列4」
        value4a5 = value5/value4/value4*10000
        if value4a5 < 18.5: # 低体重
            value4a5_a = 1
            value4a5_text = ' （低体重） '
        elif 18.5 <= value4a5 < 25: # 普通
            value4a5_a = 2
            value4a5_text = ' （普通） '
        elif 25 <= value4a5 < 30: # 肥満1
            value4a5_a = 3
            value4a5_text = ' （肥満１） '
        elif 30 <= value4a5 < 35: # 肥満2
            value4a5_a = 4
            value4a5_text = ' （肥満２） '
        elif 35 <= value4a5 < 40: # 肥満3
            value4a5_a = 5
            value4a5_text = ' （肥満３） '
        elif 40 <= value4a5: # 肥満4
            value4a5_a = 6
            value4a5_text = ' （肥満４） '
        value4a5 = format(value4a5,'.1f')
        value9 = sheet_date1.cell(row=number,column=9).value # 単位軌跡長[cm/s]
        value9 = float(value9/300)
        value9 = format(value9,'.2f')
        value10 = sheet_date1.cell(row=number,column=12).value # 矩形面積[cm2]
        value10 = float(value10/100)
        value10 = format(value10,'.2f')

        # 硬度乾燥度推定 ######################################################################
        HWlist, im_h, im_w = kaiseki_HandW.all(im, pho, value4, value5)
        kaiseki_chu.set("処理中....")
        ######################################################################################

        workbook = px.load_workbook('scanning_result_new.xlsx')
        sheet = workbook['Sheet1']

        # セルに書き込み
        sheet['V5'] = value1 # ID
        sheet['V7'] = value2 # 性別、１：男、２：女、その他
        sheet['V9'] = value3 # 年齢
        sheet['V11'] = value4 # 身長
        sheet['V13'] = value5 # 体重
        sheet['V15'] = value4a5 # BMI
        sheet['X42'] = value4a5_a # BMI
        sheet['V17'] = value6 # 接地面(%)
        sheet['V19'] = value7 # 湿潤面(%)
        sheet['V21'] = value8 # しわ(%)
        sheet['V23'] = value9 # 単位軌跡長[cm/s]
        sheet['V25'] = value10 # 矩形面積
        sheet['X27'] = HWlist[0][0] # 硬度
        sheet['X28'] = HWlist[1][0] # 硬度
        sheet['X29'] = HWlist[2][0] # 硬度
        sheet['X30'] = HWlist[3][0] # 硬度
        sheet['X31'] = HWlist[4][0] # 硬度
        sheet['X33'] = HWlist[4][1] # 水分
        sheet['X34'] = HWlist[3][1] # 水分
        sheet['X35'] = HWlist[2][1] # 水分
        sheet['X36'] = HWlist[1][1] # 水分
        sheet['X37'] = HWlist[0][1] # 水分


        #画像をエクセルに張り付ける。
        zyusyo1 = '/Users/BRLAB/others/img1/result_1angle_'+ str(numbe) +'.jpg'
        img1 = px.drawing.image.Image(zyusyo1)
        img1.width = 71 * 4
        img1.height = 26 * 15
        sheet.add_image(img1, 'F4') # 左上

        zyusyo2 = '/Users/BRLAB/others/img2/result_2press_'+ str(numbe) +'.jpg'
        img2 = px.drawing.image.Image(zyusyo2)
        img2.width = 71 * 4
        img2.height = 26 * 15
        sheet.add_image(img2, 'L4') # 左上

        zyusyo3 = '/Users/BRLAB/others/img3/result_3area_'+ str(numbe) +'.jpg'
        img3 = px.drawing.image.Image(zyusyo3)
        img3.width = 71 * 4
        img3.height = 26 * 15
        sheet.add_image(img3, 'R4') # 左上

        zyusyo4 = '/Users/BRLAB/others/img4/result_4edge_'+ str(numbe) +'.jpg'
        img4 = px.drawing.image.Image(zyusyo4)
        img4.width = 71 * 4
        img4.height = 26 * 15
        sheet.add_image(img4, 'F21') # 左上

        zyusyo5 = '/Users/BRLAB/others/img6/result_6hard_'+ str(numbe) +'.jpg'
        img5 = px.drawing.image.Image(zyusyo5)
        img5.width = 71 * 4
        img5.height = 26 * 15
        sheet.add_image(img5, 'L21') # 左上

        zyusyo6 = '/Users/BRLAB/others/img6/result_6wet_'+ str(numbe) +'.jpg'
        img6 = px.drawing.image.Image(zyusyo6)
        img6.width = 71 * 4
        img6.height = 26 * 15
        sheet.add_image(img6, 'R21') # 左上

        #エクセルファイルを保存する
        result_file_name = 'scanning_result_s1.xlsx'
        workbook.save(result_file_name)
        workbook.close()
        kaiseki_chu.set("処理完了")
        print("処理完了")


        #wwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwww


        sub_window = tk.Toplevel(baseGround) # サブ生成後にメインそのまま残すため
        sub_window.geometry("975x720") # 解析結果ウィンドウの大きさ
        sub_window.configure() # 背景色

        # メインページに足裏画像保持
        global canvas1

        img = Image.open(file_name_ans, 'r')
        w = img.width # 横幅を取得
        h = img.height # 縦幅を取得
        img = img.resize(( int(w * (545/h)), int(h * (550/h)) ))
        img = ImageTk.PhotoImage(img)
        #canvas1 = tk.Canvas(main_frame,width=500, height=600)
        canvas1.place(x = 10,y = 95)
        canvas1.create_image(240, 275, image=img)


        # 印刷ボタン押下後 _________________________________________________________
        def Print():
            print("印刷")

            open(result_file_name)
            def Print():
                win32api.ShellExecute( 0, "print", result_file_name, "/c:""%s" % win32print.GetDefaultPrinter(), ".", 0)

            workbook.active = workbook.sheetnames.index('Sheet1')

            #選択されているシートを処理対象にする
            ws = workbook.active

            #印刷範囲の設定
            ws.print_area = 'A1:Z38'

            #印刷の向きを横にする'landscape'(縦にする場合は'portrait')
            ws.page_setup.orientation = 'landscape'

            #縦１ページ、横１ページに印刷する
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.sheet_properties.pageSetUpPr.fitToPage = True

            workbook.save(result_file_name)
            Print()
            workbook.close()
            print("印刷処理完了")

            #______________________________________________________________________________


        # 印刷ボタン
        file_button = tk.Button(sub_window, text='印刷', width=8, height=1, command=Print)
        file_button.place( x = 3, y = 7)

        # 境界線
        canvas11 = tk.Canvas(sub_window,width=775,height=5)
        id11 = canvas11.create_line(0, 5, 775, 5, fill="gray0", width=1)
        canvas11.place(x = 1, y = 32 )

        # 解析後の画像
        path01 = "C:" + str(zyusyo1)# 解析後の画像住所とここを一致させてpppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppath
        img01 = Image.open(path01, 'r')
        w = img01.width # 横幅を取得
        h = img01.height # 縦幅を取得
        img01 = img01.resize(( int(w * (310/h)), int(h * (310/h)) )) # 解析結果画像表示の大きさ変える
        img01 = ImageTk.PhotoImage(img01)
        canvas111 = tk.Canvas(sub_window,width=231, height=320, bg = 'gray80')
        canvas111.place(x = 15, y = 45 )
        canvas111.create_image(117, 161, image=img01)

        path02 = "C:" + str(zyusyo2)# 解析後の画像住所とここを一致させてpppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppath
        img02 = Image.open(path02, 'r')
        w = img02.width # 横幅を取得
        h = img02.height # 縦幅を取得
        img02 = img02.resize(( int(w * (310/h)), int(h * (310/h)) )) # 解析結果画像表示の大きさ変える
        img02 = ImageTk.PhotoImage(img02)
        canvas112 = tk.Canvas(sub_window,width=231, height=320, bg = 'gray80')
        canvas112.place(x = 270,y = 45)
        canvas112.create_image(117, 161, image=img02)

        path03 = "C:" + str(zyusyo3)# 解析後の画像住所とここを一致させてpppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppath
        img03 = Image.open(path03, 'r')
        w = img03.width # 横幅を取得
        h = img03.height # 縦幅を取得
        img03 = img03.resize(( int(w * (310/h)), int(h * (310/h)) )) # 解析結果画像表示の大きさ変える
        img03 = ImageTk.PhotoImage(img03)
        canvas113 = tk.Canvas(sub_window,width=231, height=320, bg = 'gray80')
        canvas113.place( x = 525, y = 45)
        canvas113.create_image(117, 161, image=img03)

        path04 = "C:" + str(zyusyo4)# 解析後の画像住所とここを一致させてpppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppath
        img04 = Image.open(path04, 'r')
        w = img04.width # 横幅を取得
        h = img04.height # 縦幅を取得
        img04 = img04.resize(( int(w * (310/h)), int(h * (310/h)) )) # 解析結果画像表示の大きさ変える
        img04 = ImageTk.PhotoImage(img04)
        canvas114 = tk.Canvas(sub_window,width=231, height=320, bg = 'gray80')
        canvas114.place( x = 15, y = 370 )
        canvas114.create_image(117, 161, image=img04)

        path05 = "C:" + str(zyusyo5)# 解析後の画像住所とここを一致させてpppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppath
        img05 = Image.open(path05, 'r')
        w = img05.width # 横幅を取得
        h = img05.height # 縦幅を取得
        img05 = img05.resize(( int(w * (310/h)), int(h * (310/h)) )) # 解析結果画像表示の大きさ変える
        img05 = ImageTk.PhotoImage(img05)
        canvas114 = tk.Canvas(sub_window,width=231, height=320, bg = 'gray80')
        canvas114.place( x = 270, y = 370 )
        canvas114.create_image(117, 161, image=img05)

        path06 = "C:" + str(zyusyo6)# 解析後の画像住所とここを一致させてpppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppppath
        img06 = Image.open(path06, 'r')
        w = img06.width # 横幅を取得
        h = img06.height # 縦幅を取得
        img06 = img06.resize(( int(w * (310/h)), int(h * (310/h)) )) # 解析結果画像表示の大きさ変える
        img06 = ImageTk.PhotoImage(img06)
        canvas114 = tk.Canvas(sub_window,width=231, height=320, bg = 'gray80')
        canvas114.place( x = 525, y = 370 )
        canvas114.create_image(117, 161, image=img06)


        # テキストラベル
        text_ID = tk.StringVar() # 後に値を変えられるテキストラベル
        text_ID.set("【 ID 】")
        label_ID = tk.Label(sub_window,textvariable=text_ID, font=('', 12))
        label_ID.place( x = 785, y = 5)
        # テキストラベル
        text_IDx = tk.StringVar() # 後に値を変えられるテキストラベル
        text_IDx.set(str(value1))
        label_IDx = tk.Label(sub_window,textvariable=text_IDx, font=('', 12))
        label_IDx.place( x = 820, y = 25)
        # テキストラベル
        text_se = tk.StringVar() # 後に値を変えられるテキストラベル
        text_se.set("【 性別 】")
        label_se = tk.Label(sub_window,textvariable=text_se, font=('', 12))
        label_se.place( x = 785, y = 50)
        # テキストラベル
        text_sex = tk.StringVar() # 後に値を変えられるテキストラベル
        if value2 == 1:
            seibetu = '男性'
        elif value2 == 2:
            seibetu = '女性'
        else:
            seibetu = '秘密'
        text_sex.set(seibetu)
        label_sex = tk.Label(sub_window,textvariable=text_sex, font=('', 12))
        label_sex.place( x = 820, y = 70)
        # テキストラベル
        text_years = tk.StringVar() # 後に値を変えられるテキストラベル
        text_years.set("【 年齢 】")
        label_years = tk.Label(sub_window,textvariable=text_years, font=('', 12))
        label_years.place( x = 785, y = 95)
        # テキストラベル
        text_yearsx = tk.StringVar() # 後に値を変えられるテキストラベル
        text_yearsx.set(str(value3)+' 歳')
        label_yearsx = tk.Label(sub_window,textvariable=text_yearsx, font=('', 12))
        label_yearsx.place( x = 820, y = 115)
        # テキストラベル
        text_hi = tk.StringVar() # 後に値を変えられるテキストラベル
        text_hi.set('【 身長 】')
        label_hi = tk.Label(sub_window,textvariable=text_hi, font=('', 12))
        label_hi.place( x = 785, y = 140)
        # テキストラベル
        text_hix = tk.StringVar() # 後に値を変えられるテキストラベル
        text_hix.set(str(value4)+' cm')
        label_hix = tk.Label(sub_window,textvariable=text_hix, font=('', 12))
        label_hix.place( x = 820, y = 160)
        # テキストラベル
        text_we = tk.StringVar() # 後に値を変えられるテキストラベル
        text_we.set('【 体重 】')
        label_we = tk.Label(sub_window,textvariable=text_we, font=('', 12))
        label_we.place( x = 785, y = 185)
        # テキストラベル
        text_wex = tk.StringVar() # 後に値を変えられるテキストラベル
        text_wex.set(str(value5)+' kg')
        label_wex = tk.Label(sub_window,textvariable=text_wex, font=('', 12))
        label_wex.place( x = 820, y = 205)
        # テキストラベル
        text_bm = tk.StringVar() # 後に値を変えられるテキストラベル
        text_bm.set('【 BMI 】')
        label_bm = tk.Label(sub_window,textvariable=text_bm, font=('', 12))
        label_bm.place( x = 785, y = 230)
        # テキストラベル
        text_bmx = tk.StringVar() # 後に値を変えられるテキストラベル
        text_bmx.set(str(value4a5)+value4a5_text)
        label_bmx = tk.Label(sub_window,textvariable=text_bmx, font=('', 12))
        label_bmx.place( x = 820, y = 250)
        # テキストラベル
        text_co = tk.StringVar() # 後に値を変えられるテキストラベル
        text_co.set('【 接地面積比率 】')
        label_co = tk.Label(sub_window,textvariable=text_co, font=('', 12))
        label_co.place( x = 785, y = 275)
        # テキストラベル
        text_cox = tk.StringVar() # 後に値を変えられるテキストラベル
        text_cox.set(str(value6)+' %')
        label_cox = tk.Label(sub_window,textvariable=text_cox, font=('', 12))
        label_cox.place( x = 820, y = 295)
        # テキストラベル
        text_ed = tk.StringVar() # 後に値を変えられるテキストラベル
        text_ed.set('【 しわ・ひび割れ比率 】')
        label_ed = tk.Label(sub_window,textvariable=text_ed, font=('', 12))
        label_ed.place( x = 785, y = 320)
        # テキストラベル
        text_edx = tk.StringVar() # 後に値を変えられるテキストラベル
        text_edx.set(str(value8)+' %')
        label_edx = tk.Label(sub_window,textvariable=text_edx, font=('', 12))
        label_edx.place( x = 820, y = 340)
        # テキストラベル
        text_zs = tk.StringVar() # 後に値を変えられるテキストラベル
        text_zs.set('【 重心動揺/単位軌跡長 】')
        label_zs = tk.Label(sub_window,textvariable=text_zs, font=('', 12))
        label_zs.place( x = 785, y = 365)
        # テキストラベル
        text_zsx = tk.StringVar() # 後に値を変えられるテキストラベル
        text_zsx.set(str(value9)+' cm/s')
        label_zsx = tk.Label(sub_window,textvariable=text_zsx, font=('', 12))
        label_zsx.place( x = 820, y = 385)
        # テキストラベル
        text_zk = tk.StringVar() # 後に値を変えられるテキストラベル
        text_zk.set('【 重心動揺/矩形面積 】')
        label_zk = tk.Label(sub_window,textvariable=text_zk, font=('', 12))
        label_zk.place( x = 785, y = 410)
        # テキストラベル
        text_zkx = tk.StringVar() # 後に値を変えられるテキストラベル
        text_zkx.set(str(value10)+' cm2')
        label_zkx = tk.Label(sub_window,textvariable=text_zkx, font=('', 12))
        label_zkx.place( x = 820, y = 430)
        # テキストラベル
        text_h0 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_h0.set('【 硬度 】')
        label_h0 = tk.Label(sub_window,textvariable=text_h0, font=('', 12))
        label_h0.place( x = 785, y = 455)
        # テキストラベル
        text_h1 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_h1.set('軟質 ： '+str(HWlist[0][0])+' %')
        label_h1 = tk.Label(sub_window,textvariable=text_h1, font=('', 12))
        label_h1.place( x = 865, y = 455)
        # テキストラベル１
        text_h2 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_h2.set('弱軟 ： '+str(HWlist[1][0])+' %')
        label_h2 = tk.Label(sub_window,textvariable=text_h2, font=('', 12))
        label_h2.place( x = 865, y = 475)
        # テキストラベル
        text_h3 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_h3.set('普通 ： '+str(HWlist[2][0])+' %')
        label_h3 = tk.Label(sub_window,textvariable=text_h3, font=('', 12))
        label_h3.place( x = 865, y = 495)
        # テキストラベル
        text_h4 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_h4.set('弱硬 ： '+str(HWlist[3][0])+' %')
        label_h4 = tk.Label(sub_window,textvariable=text_h4, font=('', 12))
        label_h4.place( x = 865, y = 515)
        # テキストラベル
        text_h5 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_h5.set('硬質 ： '+str(HWlist[4][0])+' %')
        label_h5 = tk.Label(sub_window,textvariable=text_h5, font=('', 12))
        label_h5.place( x = 865, y = 535)
        # テキストラベル
        text_w0 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_w0.set('【 水分量 】')
        label_w0 = tk.Label(sub_window,textvariable=text_w0, font=('', 12))
        label_w0.place( x = 785, y = 560)
        # テキストラベル
        text_w1 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_w1.set('湿潤 ： '+str(HWlist[4][1])+' %')
        label_w1 = tk.Label(sub_window,textvariable=text_w1, font=('', 12))
        label_w1.place( x = 865, y = 560)
        # テキストラベル１
        text_w2 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_w2.set('弱潤 ： '+str(HWlist[3][1])+' %')
        label_w2 = tk.Label(sub_window,textvariable=text_w2, font=('', 12))
        label_w2.place( x = 865, y = 580)
        # テキストラベル
        text_w3 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_w3.set('普通 ： '+str(HWlist[2][1])+' %')
        label_w3 = tk.Label(sub_window,textvariable=text_w3, font=('', 12))
        label_w3.place( x = 865, y = 600)
        # テキストラベル
        text_w4 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_w4.set('弱乾 ： '+str(HWlist[1][1])+' %')
        label_w4 = tk.Label(sub_window,textvariable=text_w4, font=('', 12))
        label_w4.place( x = 865, y = 620)
        # テキストラベル
        text_w5 = tk.StringVar() # 後に値を変えられるテキストラベル
        text_w5.set('乾燥 ： '+str(HWlist[0][1])+' %')
        label_w5 = tk.Label(sub_window,textvariable=text_w5, font=('', 12))
        label_w5.place( x = 865, y = 640)


    #*********************************************************************************************
    global datekey
    global im
    global imgkai
    global resolt
    global file_name_ans

    # 解析ボタン
    main_button = tk.Button(main_frame, text="解析",width=8, height=1, command=kaiseki)
    main_button.place( x = 3, y = 31 )

    # # 配置的にここにテキストラベル2 投下
    label2.place( x = 80, y = 31)

    # 境界線
    canvas = tk.Canvas(main_frame,width=1400,height=5)
    id = canvas.create_line(0, 5, 1398, 5, fill="gray0", width=1)
    canvas.place( x = 1, y = 57)


    # 配置的にここにテキストラベル投下
    label.place( x = 3, y = 64)



# はじめのGUI いくつかの条件のボタンを選択し、その後 id() へ(被験者番号入力)
def start_jouken():

    frame_start = tk.Frame(baseGround, width=500, height=700)
    frame_start.pack()
    def click_btnStart1(): # 下のスタートボタン1を押した後のイベント
        global jnum
        jnum=1
        frame_start.destroy()
        id()
    def click_btnStart2(): # 下のスタートボタン2を押した後のイベント
        global jnum
        jnum=2
        frame_start.destroy()
        id()
    def click_btnStart3(): # 下のスタートボタン3を押した後のイベント
        global jnum
        jnum=3
        frame_start.destroy()
        id()
    def click_btnStart4(): # 下のスタートボタン4を押した後のイベント
        global jnum
        jnum=4
        frame_start.destroy()
        id()
    def file_dialog(): # 下のスタートボタン5を押した後のイベント
        frame_start.destroy()
        kaiseki_gui()

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_start, text="まだ台に乗らないでください",font=("",30))
    label2.place(x=20, y=10)
    label2['fg']="red"

    buttonStart1 = tk.Button(frame_start,text = '脱いですぐSTART', command=click_btnStart1,font=("",20))
    buttonStart1.place(x=50, y=145, relwidth=0.75, relheight= 0.125)

    buttonStart2 = tk.Button(frame_start,text = '脱いでしばらくしてSTART', command=click_btnStart2,font=("",20))
    buttonStart2.place(x=50, y=245, relwidth=0.75, relheight= 0.125)

    buttonStart3 = tk.Button(frame_start,text = '+20kgでSTART', command=click_btnStart3,font=("",20))
    buttonStart3.place(x=50, y=345, relwidth=0.75, relheight= 0.125)

    buttonStart4 = tk.Button(frame_start,text = 'ウエットティッシュでSTART', command=click_btnStart4,font=("",20))
    buttonStart4.place(x=50, y=445, relwidth=0.75, relheight= 0.125)

    buttonStart5 = tk.Button(frame_start,text = '画像処理/解析', command=file_dialog,font=("",20))
    buttonStart5.place(x=50, y=545, relwidth=0.75, relheight= 0.125)

# スキャンスタート！！終わったら重心動揺測定へ ######################################################################
def scan():
    def click_btnFinish():
        frame_scan.destroy()
        jushin()

    def click_btnEnter():
        thread1 = threading.Thread(target=start_scan)
        thread1.start()

    def start_scan():
        global file_cope
        ser.write(b"3") # シリアル通信でArduino に case3 のスキャン命令
        buttonEnter.destroy()
        # buttonFinish.destroy()
        label3 = tk.Label(frame_scan, text='スキャン中', background='lightgray',font=("",25))
        label3.place(x=50, y=90, width=400, height=105)
        exitnum=10
        # exitnum=11
        while exitnum==10:
            line2 = int(ser.readline())
            if line2 !=11:
                pass
            else:
                exitnum=11
        label.destroy()
        label2.destroy()
        label3.destroy()

        list_of_files = glob.glob('C:/Users/BRLAB/Pictures/*.jpg') # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime) # key=os.path.getctime は更新日時、
        karinumber = latest_file.removeprefix('C:/Users/BRLAB/Pictures\img')
        karinumber = karinumber.removesuffix('.jpg')

        print(latest_file)
        print(enumber-1)
        print(jnum)
        print(idnum)
        print(seibetsu)
        print(nenrei)
        print(shincho)


        rename_file='C:/Users/BRLAB/Desktop/photo/'+ str(enumber-1)+'_'+str(jnum)+'_'+str(idnum)+'_'+str(today)+'.jpg'
        file_cope = str(enumber-1)+'_'+str(jnum)+'_'+str(idnum)+'_'+str(today)
        os.rename(latest_file,rename_file) # os.rename(変更前ファイル、変更後ファイル)

        label4 = tk.Label(frame_scan, text="スキャンが完了しました\n終了を押してください",font=("",30))
        label4.place(x=50, y=0)

        label5 = tk.Label(frame_scan, text='スキャン完了', background='lightgray',font=("",25))
        label5.place(x=50, y=90, width=400, height=105)

        buttonFinish = tk.Button(frame_scan,text = '重心測定へ', command=click_btnFinish)
        buttonFinish.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

    frame_scan = tk.Frame(baseGround, width=500, height=600)
    frame_scan.pack()

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_scan, text="足裏をスキャンします\nスキャンを押してください",font=("",30))
    label2.place(x=50, y=0)

    label = tk.Label(frame_scan, text='スキャン準備完了', background='lightgray',font=("",25))
    label.place(x=50, y=90, width=400, height=105)

    buttonEnter = tk.Button(frame_scan,text = 'スキャン', command=click_btnEnter,font=("",30),fg='red')
    buttonEnter.place(x=50, y=205, relwidth=0.36, relheight= 0.125)
    # buttonFinish = tk.Button(frame_scan,text = '終了', command=click_btnFinish)
    # buttonFinish.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

# 体重測定のためのキャリブレーション ################################################################
def taijucali():
    ser.write(b"1") # シリアル通信でArduino に case1 を指令（8秒間）
    frame_cali = tk.Frame(baseGround, width=500, height=600)
    frame_cali.pack()
    def timer():
        for i in range(8):
            label['text']=""+str(7-i)+"秒"
            time.sleep(1)
        frame_cali.destroy()
        taiju()

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_cali, text="まだ台に乗らないでください\n処理中です\n残り時間",font=("",30))
    label2.place(x=30, y=0)
    label2['fg']="red"

    label = tk.Label(frame_cali, text='0', background='lightgray',font=("",50))
    label.place(x=50, y=130, width=400, height=105)

    thread3 = threading.Thread(target=timer)
    thread3.start()

# 体重測定スタート！！ ######################################################################
def taiju():
    def click_btnFinish():
        ser.write(b"9") # case1 内の体重測定コール（'9'）
        # ser.close()
        frame_kg.destroy()
        taijucali()

    def click_btnEnter(): # 決定ボタンが押された後のイベント（スキャンへ）
        frame_kg.destroy()
        scan()

    def taijusave():
        global weight
        weight=0
        weightlist=[0,0]
        exitnum=0
        while exitnum==0:
            line = str(ser.readline()) # readlineはファイルから1行だけ読み出す。readlinesはファイルの内容を全て読み出し、1行ごとのリストにします。
            weightlist.append(int(line.split(',')[1])/10)
            weight = int(line.split(',')[1])/10
            print(str(weight))
            label['text'] = ""+str(weight)+"kg"
            if int(line.split(',')[1])>300:
                if max(weightlist[-15:-1])-min(weightlist[-15:])<2:
                    weight=sum(weightlist[-15:])/15
                    weight=round(weight,1)
                    print(weight)
                    exitnum=1
                    ser.write(b"9")
                else:
                    pass
            else:
                pass

        ws = wb.active
        enumber = 1
        while not ws.cell(enumber, 1).value is None:
            enumber += 1
        enumber -= 1
        ws.cell(enumber, 8).value = float(weight)
        wb.save('hikensya.xlsx')
        print(weight)

        label['text'] = str(weight)
        weight=label['text']
        print(weight)

        buttonEnter = tk.Button(frame_kg,text = '決定', command=click_btnEnter,font=("",50))
        buttonEnter.place(x=50, y=205, relwidth=0.36, relheight= 0.125)
        label2.destroy()
        label3 = tk.Label(frame_kg, text="決定を押してください",font=("",30),fg="red")
        label3.place(x=40, y=0)


    frame_kg = tk.Frame(baseGround, width=500, height=600)
    frame_kg.pack()

    # ラベルやボタンの作成と配置

    label2 = tk.Label(frame_kg, text="台に乗ってください",font=("",30))
    label2.place(x=60, y=0)

    label = tk.Label(frame_kg, text='0', background='lightgray',font=("",50))
    label.place(x=50, y=50, width=400, height=105)

    buttonFinish = tk.Button(frame_kg,text = '戻る', command=click_btnFinish)
    buttonFinish.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

    thread2 = threading.Thread(target=taijusave)
    thread2.start()

def jushin(): # 重心動揺測定！！ ##############################################################
    frame_jushin = tk.Frame(baseGround, width=500, height=600)
    frame_jushin.pack()
    def click_btnEnter():
        frame_jushin.destroy()
        finish_sokutei()

    def timer():
      for i in range(11):
        label['text']=""+str(10-i)+"秒"
        time.sleep(1)
      thread4 = threading.Thread(target=jushin_sokutei)
      thread4.start()
      for i in range(36):
        label['text']=""+str(35-i)+"秒"
        time.sleep(1)
      ser.write(b"9")
      buttonEnter = tk.Button(frame_jushin,text = '決定', command=click_btnEnter,font=("",50))
      buttonEnter.place(x=50, y=250, relwidth=0.36, relheight= 0.125)
      buttonEnter['fg']="red"
      label4 = tk.Label(frame_jushin, text="重心計測が終了しました\n決定を押してください",font=("",30))
      label4.place(x=55, y=0)
      label4['bg']="gray94"

    def jushin_sokutei():
        global jushinlist
        ser.write(b"4") # Arduino にケース４を命令、重心測定
        label2.destroy()
        label3 = tk.Label(frame_jushin, text="重心を計測しています\n終了まで",font=("",30))
        label3.place(x=50, y=0)
        label3['fg']="black"
        enumber=1

        ws = wb.active
        while not ws.cell(1, enumber).value is None:
            enumber =enumber + 1
        enumber=enumber-1

        jushinlist = [[0 for i in range(3000)] for j in range(2)]

        for i in range(3000):
            line = str(ser.readline())
            print(line)
            try:
                jushinlist[0][i]=int(line.split(',')[1]) # xposition
                jushinlist[1][i]=int(line.split(',')[2]) # yposition
                # Arduino でweight1~4を使い重心の(pointX,pointY)を計算している。それをser.を使って読み込んでいる。
            except:
                jushinlist[0][i]=99999
                jushinlist[1][i]=99999

        filename = 'C:/Users/BRLAB/Desktop/jushin/'+ str(file_cope) +'.csv'
        with open(str(filename), 'a',newline="") as csv_file:
            fieldnames = ['time','xposition', 'yposition']
            writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
            writer.writeheader()
            for i in range(3000):
                writer.writerow({'time': i,'xposition':  jushinlist[0][i], 'yposition': jushinlist[1][i]})


        val=2

        #ローパスフィルタ 今回使用
        def lowpass(x, samplerate, fp, fs, gpass, gstop):
            fn = samplerate / 2                           #ナイキスト周波数
            wp = fp / fn                                  #ナイキスト周波数で通過域端周波数を正規化
            ws = fs / fn                                  #ナイキスト周波数で阻止域端周波数を正規化
            N, Wn = signal.buttord(wp, ws, gpass, gstop)  #オーダーとバターワースの正規化周波数を計算
            b, a = signal.butter(N, Wn, "low")            #フィルタ伝達関数の分子と分母を計算
            y = signal.filtfilt(b, a, x)                  #信号に対してフィルタをかける
            return y

        # csvから列方向に順次フィルタ処理を行い保存する関数
        def csv_filter(in_file, out_file, type):
            df = pd.read_csv(in_file, encoding='SHIFT-JIS')                  # ファイル読み込み
            dt = 1/80                                              # 時間刻み

            # データフレームを初期化
            df_filter = pd.DataFrame()
            df_filter[df.columns[0]] = df.T.iloc[0]

            # ローパスの設定-----------------------------------------------------------------------------
            # fp_lp = 15                                                       # 通過域端周波数[Hz]
            # fs_lp = 30                                                       # 阻止域端周波数[Hz]

            fp_lp = 3                                                      # 通過域端周波数[Hz]
            fs_lp = 5                                                       # 阻止域端周波数[Hz]

            gpass = 3                                                        # 通過域端最大損失[dB]
            gstop = 40                                                       # 阻止域端最小損失[dB]

            # 列方向に順次フィルタ処理をするコード
            for i in range(len(df.T)-1):
                data = df.T.iloc[i+1]                       # フィルタ処理するデータ列を抽出

                # フィルタ処理の種類を文字列で読み取って適切な関数を選択する
                if type == 'lp':
                    # ローパスフィルタを実行
                    # print('wave=', i, ':Lowpass.')
                    data = lowpass(x=data, samplerate=1 / dt,
                                fp=fp_lp, fs=fs_lp,
                                gpass=gpass, gstop=gstop)
                else:
                    # 文字列が当てはまらない時はパス（動作テストでフィルタかけたくない時はNoneとか書いて実行するとよい）
                    pass

                data = pd.Series(data)                                       # dataをPandasシリーズデータへ変換
                df_filter[df.columns[i + 1] + '_filter'] = data              # 保存用にデータフレームへdataを追加

            df_filter.to_csv(out_file)                                       # フィルタ処理の結果をcsvに保存

            return df, df_filter

        # p=filename
        # p=str(p.split('.')[0])
        # print(p)
        df, df_filter = csv_filter(in_file='C:/Users/BRLAB/Desktop/jushin/'+str(file_cope)+'.csv', out_file='C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv', type='lp')
        # ws.cell(val, 1).value = p
        #print(ws.cell(1, 2).value) (1,b)
        with open('C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv', 'r') as csvfile:
        # with open('101_1_0_20211204.csv', 'r') as csvfile:
            # print(p)
            reader = csv.reader(csvfile)
            csv_data = list(reader)
            # sum=((int(csv_data[1][1])-int(csv_data[2][1]))**2+(int(csv_data[1][2])-int(csv_data[2][2]))**2)**0.5
            # print(sum)
            # print(csv_data[1][1])
            # print(csv_data[1][2])
            sum_all=0
            sum_all=float(sum_all)
            sum_x=0
            sum_x=float(sum_x)
            sum_y=0
            sum_y=float(sum_y)
            df = pd.read_csv('C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv',index_col=0)
            max_x=df.max()['xposition_filter']
            min_x=df.min()['xposition_filter']
            maxd_x=max_x-min_x
            max_y=df.max()['yposition_filter']
            min_y=df.min()['yposition_filter']
            maxd_y=max_y-min_y
            for i in range(2400):
                sum_all=sum_all+((float(csv_data[i+1][2])-float(csv_data[i+2][2]))**2+(float(csv_data[i+1][3])-float(csv_data[i+2][3]))**2)**0.5
                sum_x=sum_x+abs(float(csv_data[i+1][2])-float(csv_data[i+2][2]))
                sum_y=sum_y+abs(float(csv_data[i+1][3])-float(csv_data[i+2][3]))
            # print(sum_all)
        # ws.cell(val, 6).value = str(sum_all)
        # ws.cell(val, 7).value = str(sum_x)
        # ws.cell(val, 8).value = str(sum_y)
        # ws.cell(val, 9).value = str(maxd_x)
        # ws.cell(val, 10).value = str(maxd_y)
        #print(ws.cell(1, 2).value) (1,b)
        # wb.save('jushin_kekka.xlsx')
        val=val+1

        fig = plt.figure(figsize=[15,15])

        input_csv = pd.read_csv('C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv')
        first_column_data = input_csv[input_csv.keys()[2]]
        second_column_data = input_csv[input_csv.keys()[3]]

        plt.xlabel(input_csv.keys()[2],fontsize=20)
        plt.ylabel(input_csv.keys()[3],fontsize=20)

        plt.plot(first_column_data, second_column_data, linestyle='solid', marker='')
        plt.tick_params(labelsize=20)
        #plt.show()
        fig.savefig('C:/Users/BRLAB/Desktop/jushin/csv_filter_photo/'+str(file_cope)+"_filter.png")


        print("総軌跡長＝",sum_all)
        print("x最大変位=",maxd_x)
        print("y最大変位＝",maxd_y)

        xy_area = maxd_x*maxd_y
        print("矩形面積＝",xy_area)

        ws = wb.active
        enumber = 1
        while not ws.cell(enumber, 1).value is None:
            enumber += 1
        enumber -= 1
        ws.cell(enumber, 9).value = float(sum_all)
        ws.cell(enumber, 10).value = float(maxd_x)
        ws.cell(enumber, 11).value = float(maxd_y)
        ws.cell(enumber, 12).value = float(xy_area)


        wb.save('/Users/BRLAB/hikensya.xlsx')

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_jushin, text="重心を35秒間測定します\n直立を維持してください\n開始まで",font=("",30))
    label2.place(x=30, y=0)
    label2['fg']="black"

    label = tk.Label(frame_jushin, text='0', background='lightgray',font=("",50))
    label.place(x=50, y=130, width=400, height=105)


    thread3 = threading.Thread(target=timer)
    thread3.start()

# ここから「診断2 」入力 #####################################################################################
def shindan2():
    def click_btn1(): # 'はい' が押されたとき
        if label['text'] == 'いいえ':
            label['text'] = 'はい'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        elif label['text'] == '不明':
            label['text'] = 'はい'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = 'はい'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btn2(): # 'いいえ' が押されたとき
        if label['text'] == 'はい':
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        elif label['text'] == '不明':
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btn3(): # '不明' が押されたとき
        if label['text'] == 'はい':
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        elif label['text'] == 'いいえ':
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btnFinish(): # 「中断」ボタンを押されたときのイベント（スタート画面へ）
        frame_shindan2.destroy()
        finish_sokutei()
    def click_btnEnter(): # 「決定」ボタンが押されたときのイベント（はい/いいえ/不明を番号でエクセルに記録、その後shindan2()へ）
        global komoku2
        komoku2=0
        if label['text'] == 'はい':
            komoku2=1
            ws = wb.active # エクセルをアクティブに
            ws.cell(enumber, 14).value = komoku2 # 指定のセルに書き込み
            wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan2.destroy()
            finish_sokutei() # 測定終了画面へ、診断項目を増やしたければ、上に新たなshindan3 でも作って、ここをshindan3 と書けばそこに飛ぶ
        elif label['text'] == 'いいえ':
            komoku2=2
            ws = wb.active # エクセルをアクティブに
            ws.cell(enumber, 14).value = komoku2 # 指定のセルに書き込み
            wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan2.destroy()
            finish_sokutei() # 測定終了画面へ、診断項目を増やしたければ、上に新たなshindan3 でも作って、ここをshindan3 と書けばそこに飛ぶ
        elif label['text'] == '不明':
            komoku2=3
            ws = wb.active # エクセルをアクティブに
            ws.cell(enumber, 14).value = komoku2 # 指定のセルに書き込み
            wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan2.destroy()
            finish_sokutei() # 測定終了画面へ、診断項目を増やしたければ、上に新たなshindan3 でも作って、ここをshindan3 と書けばそこに飛ぶ
        else:
            label2['fg']="red"
    frame_shindan2 = tk.Frame(baseGround, width=600, height=600)
    frame_shindan2.pack()
    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_shindan2, text="現在、糖尿病神経障害罹患者ですか？",font=("",20))
    label2.place(x=50, y=0) # 配置
    label = tk.Label(frame_shindan2, text='', background='lightgray',font=("",30))
    label.place(x=100, y=50, width=400, height=105) # 配置
    button1 = tk.Button(frame_shindan2,text = 'はい', command=click_btn1,font=("",20))
    button1.place(x=50, y=205, relwidth=0.16, relheight= 0.125) # 配置
    button2 = tk.Button(frame_shindan2,text = 'いいえ', command=click_btn2,font=("",20))
    button2.place(x=153, y=205, relwidth=0.16, relheight= 0.125) # 配置
    button3 = tk.Button(frame_shindan2,text = '不明', command=click_btn3,font=("",20))
    button3.place(x=256, y=205, relwidth=0.16, relheight= 0.125) # 配置
    buttonEnter = tk.Button(frame_shindan2,text = '決定', command=click_btnEnter,font=("",30))
    buttonEnter.place(x=359, y=205, relwidth=0.32, relheight= 0.125) # 配置
    buttonFinish = tk.Button(frame_shindan2,text = '中断', command=click_btnFinish)
    buttonFinish.place(x=455, y=295, relwidth=0.16, relheight= 0.125) # 配置

# ここから「診断１」入力 #####################################################################################
def shindan1():
    def click_btn1(): # 'はい' が押されたとき
        if label['text'] == 'いいえ':  # 'はい' が押されたときにlabelを"はい"にする。
            label['text'] = 'はい'     # label は"はい" "いいえ" "不明" の返答結果を表示するラベル
            buttonEnter['fg'] = "red" # buttonEnter は「決定」ボタン
            label2['fg']="black"      # label2 は「現在、糖尿病罹患者ですか？」の文のラベル
        elif label['text'] == '不明':
            label['text'] = 'はい'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = 'はい'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btn2(): # 'いいえ' が押されたとき
        if label['text'] == 'はい': # 'いいえ' が押されたときにlabelを"いいえ"にする。
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        elif label['text'] == '不明':
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btn3(): # '不明' が押されたとき
        if label['text'] == 'はい': # '不明' が押されたときにlabelを"不明"にする。
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        elif label['text'] == 'いいえ':
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btnFinish(): # 「中断」ボタンを押されたときのイベント（スタート画面へ戻る）
        frame_shindan1.destroy() # frame_shindan1 画面を消す
        finish_sokutei() # 1794行目あたりの finish_sokutei を実行
    def click_btnEnter(): # 「決定」ボタンが押されたときのイベント（はい/いいえ/不明を番号でエクセルに記録、その後shindan2()へ）
        global komoku1
        komoku1=0 # 項目1 の結果をエクセルに送る用の番号
        if label['text'] == 'はい':
            komoku1=1
            ws = wb.active # 記録用エクセルファイルをアクティブに
            ws.cell(enumber, 13).value = komoku1 # 指定セルに書き込み
            wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan1.destroy() # frame_shindan1 画面を消す
            shindan2() # 診断2 へ
        elif label['text'] == 'いいえ':
            komoku1=2
            ws = wb.active # エクセルをアクティブに
            ws.cell(enumber, 13).value = komoku1 # 指定セルに書き込み
            wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan1.destroy()
            shindan2() # 診断2 へ
        elif label['text'] == '不明':
            komoku1=3
            ws = wb.active # エクセルをアクティブに
            ws.cell(enumber, 13).value = komoku1 # 指定セルに書き込み
            wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan1.destroy()
            shindan2() # 診断2 へ
        else:
            label2['fg']="red"
    frame_shindan1 = tk.Frame(baseGround, width=600, height=600)
    frame_shindan1.pack()
    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_shindan1, text="現在、糖尿病罹患者ですか？",font=("",20))
    label2.place(x=50, y=0) # 配置
    label = tk.Label(frame_shindan1, text='', background='lightgray',font=("",30))
    label.place(x=100, y=50, width=400, height=105) # 配置
    button1 = tk.Button(frame_shindan1,text = 'はい', command=click_btn1,font=("",20))
    button1.place(x=50, y=205, relwidth=0.16, relheight= 0.125) # 配置
    button2 = tk.Button(frame_shindan1,text = 'いいえ', command=click_btn2,font=("",20))
    button2.place(x=153, y=205, relwidth=0.16, relheight= 0.125) # 配置
    button3 = tk.Button(frame_shindan1,text = '不明', command=click_btn3,font=("",20))
    button3.place(x=256, y=205, relwidth=0.16, relheight= 0.125) # 配置
    buttonEnter = tk.Button(frame_shindan1,text = '決定', command=click_btnEnter,font=("",30))
    buttonEnter.place(x=359, y=205, relwidth=0.32, relheight= 0.125) # 配置
    buttonFinish = tk.Button(frame_shindan1,text = '中断', command=click_btnFinish)
    buttonFinish.place(x=455, y=295, relwidth=0.16, relheight= 0.125) # 配置


# 重心動揺測定終了後、「はじめに戻る」もしくは「画像処理/解析」のボタン表示 ###########################################################3
def finish_sokutei(): # 測定終了画面
    frame_fini = tk.Frame(baseGround, width=500, height=600) # 測定終了画面、フレーム、
    frame_fini.pack() # 配置、特に指定なし
    def click_btnStart1():   # buttonStart1 が押されたときのイベント
        global jnum
        jnum=1
        frame_fini.destroy()  # 測定終了画面を消す
        start_jouken()        # start_jouken を実行
    def click_btnStart2(): # buttonStart2 が押されたときのイベント
        frame_fini.destroy()  # 測定終了画面を消す
        kaiseki_gui()         # kaiseki_guiを実行
    def click_btnStart3(): # buttonStart3 が押されたときのイベント
        frame_fini.destroy()  # 測定終了画面を消す
        shindan1()            # shindan1 を実行
    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_fini, text="計測終了です。",font=("",30))
    label2.place(x=120, y=100)
    label2['fg']="black"

    buttonStart1 = tk.Button(frame_fini,text = 'はじめに戻る', command=click_btnStart1,font=("",20))
    buttonStart1.place(x=50, y=195, relwidth=0.75, relheight= 0.125)

    buttonStart2 = tk.Button(frame_fini,text = '画像処理/解析', command=click_btnStart2,font=("",20))
    buttonStart2.place(x=50, y=295, relwidth=0.75, relheight= 0.125)

    buttonStart3 = tk.Button(frame_fini,text = '診断', command=click_btnStart3,font=("",20))
    buttonStart3.place(x=50, y=395, relwidth=0.75, relheight= 0.125)




# メインウィンドウを作成

baseGround = tk.Tk()
# ウィンドウのサイズを設定
#baseGround.geometry('500x600')
baseGround.state('zoomed')
# 画面タイトル
baseGround.title('足裏測定')

start_jouken()
# scan()

baseGround.mainloop()
