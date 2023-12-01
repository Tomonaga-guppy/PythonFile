import tkinter as tk
import openpyxl as px  # excel 書き込み用、pip install openpyxl 必要
import datetime
import time
import threading

examinee_xlsx = 'C:/Users/zutom/.vscode/PythonFile/FootScan/Practice/examinee.xlsx'
wb = px.load_workbook(examinee_xlsx) # 被験者情報の保存先のエクセルファイル
ws = wb.active

def step1_start():

    frame_start = tk.Frame(baseGround, width=500, height=700)
    frame_start.pack()
    def click_btnStart1(): # 下のスタートボタン1を押した後のイベント
        global jnum
        jnum=1
        frame_start.destroy()
        step2_input()

    def file_dialog(): # 下のスタートボタン5を押した後のイベント
        frame_start.destroy()
        kaiseki_gui()

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_start, text="まだ台に乗らないでください",font=("",30))
    label2.place(x=20, y=10)
    label2['fg']="red"

    buttonStart1 = tk.Button(frame_start,text = '開始', command=click_btnStart1,font=("",20))
    buttonStart1.place(x=50, y=200, relwidth=0.75, relheight= 0.125)

    buttonStart5 = tk.Button(frame_start,text = '画像処理/解析', command=file_dialog,font=("",20))
    buttonStart5.place(x=50, y=400, relwidth=0.75, relheight= 0.125)

def step2_input():
    frame_input = tk.Frame(baseGround, width=800, height=600)
    frame_input.pack()
    global root, txtEntry

    def click_man():
        global gender
        gender = int(0)
        push_man.config(fg='#0000ff')
        push_woman.config(fg='#DCDCDC')

    def click_woman():
        global gender
        gender = int(1)
        push_man.config(fg='#DCDCDC')
        push_woman.config(fg='#ff0000')

    def click_btnEnter(): # 決定ボタンで、被験者情報をエクセルに書き込み、体重測定開始
        global today
        global height
        global enumber

        enumber=1
        idnum = txt_id.get()
        age = txt_age.get()
        height = txt_height.get()

        dt_now = datetime.datetime.now() # 今日の日付
        year=dt_now.year
        month=dt_now.month
        month = '{0:0>2}'.format(month)
        day =dt_now.day
        day = '{0:0>2}'.format(day)
        hour = dt_now.hour
        minute = dt_now.minute
        second = dt_now.second
        today=str(year)+'/'+str(month)+'/'+str(day)+'/'+str(hour)+':'+str(minute)+':'+str(second)

        while not ws.cell(enumber, 1).value is None: # exminee.xlsx でデータの無い行まで１からプラス１していく（必然と最後尾に付く）
            enumber += 1

        ws.cell(enumber, 1).value = int(enumber)-1 # ここから下、被験者情報をエクセルに保存、cell(enumber, 1)はenumber行 1 列目のセルを指定
        ws.cell(enumber, 2).value = today
        # ws.cell(enumber, 3).value = int(condition_num)
        ws.cell(enumber, 4).value = idnum
        ws.cell(enumber, 5).value = gender
        ws.cell(enumber, 6).value = int(age)
        ws.cell(enumber, 7).value = int(height)
        #ws.cell(enumber, 8).value = float(weight)
        wb.save(examinee_xlsx)
        label_end = tk.Label(frame_input,text="保存完了",bg = '#ccccff',font=('',20))
        label_end.place(x=500, y=520, relwidth=0.3, relheight= 0.1)

        frame_input.destroy()
        step3_BodyWeightCali()

    def click_skip():
        frame_input.destroy()
        step3_BodyWeightCali()

    #ラベル
    label_explanation = tk.Label(frame_input,text='以下の項目を入力してください',bg = '#ccccff',font=('',20))
    label_gender = tk.Label(frame_input,text='性別',bg = '#ccccff',font=('',20))
    label_age = tk.Label(frame_input,text='年齢',bg = '#ccccff',font=('',20))
    label_id = tk.Label(frame_input,text='氏名（カナ）',bg = '#ccccff',font=('',20))
    label_height = tk.Label(frame_input,text='身長 (cm)',bg = '#ccccff',font=('',20))

    #テキストボックス
    txt_id = tk.Entry(frame_input,font=('',20),width=20)
    txt_age = tk.Entry(frame_input,font=('',20),width=20)
    txt_height = tk.Entry(frame_input,font=('',20),width=20)

    #ボタン
    push_man = tk.Button(frame_input,text = '男性', command=click_man,font=("",20),fg='#0000ff')
    push_woman = tk.Button(frame_input,text = '女性', command=click_woman,font=("",20),fg='#ff0000')
    buttonEnter = tk.Button(frame_input,text = '決定', command=click_btnEnter,font=("",20))

    label_explanation.place(x=400,y=20,anchor='n')
    label_gender.place(x=150,y=100)
    push_man.place(x=200, y=150,height=40)
    push_woman.place(x=300, y=150,height=40)
    label_age.place(x=150,y=200)
    txt_age.place(x=200, y=250,height=40)
    label_id.place(x=150,y=300)
    txt_id.place(x=200, y=350,height=40)
    label_height.place(x=150,y=400)
    txt_height.place(x=200, y=450,height=40)
    buttonEnter.place(x=350, y=520, relwidth=0.15, relheight= 0.1)

    buttonSkip = tk.Button(frame_input,text = 'スキップ', command=click_skip,font=("",20))
    buttonSkip.place(x=500, y=520, relwidth=0.15, relheight= 0.1)

def step3_BodyWeightCali():

    frame_cali = tk.Frame(baseGround, width=800, height=600)
    frame_cali.pack()

    def timer():
        for i in range(8):
            label['text']=""+str(7-i)+"秒"
            time.sleep(0.5)
        frame_cali.destroy()
        step4_BodyWeight()

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_cali, text="まだ台に乗らないでください\n処理中です\n残り時間",font=("",30))
    label2.place(x=30, y=0)
    label2['fg']="red"

    label = tk.Label(frame_cali, text='0', background='lightgray',font=("",50))
    label.place(x=50, y=130, width=400, height=105)
    thread3 = threading.Thread(target=timer)
    thread3.start()

def step4_BodyWeight():
    def click_btnFinish():
        # ser.write(b"9") # case1 内の体重測定コール（'9'）
        # ser.close()
        frame_kg.destroy()
        step3_BodyWeightCali()

    def click_btnEnter(): # 決定ボタンが押された後のイベント（スキャンへ）
        frame_kg.destroy()
        step5_scan()

    def BodyWeightSave():
        global weight
        weight=0
        weightlist=[0,0]
        exitnum=0
        while exitnum==0:
            line = str(ser.readline()) # readlineはファイルから1行だけ読み出す。readlinesはファイルの内容を全て読み出し、1行ごとのリストにします。
            weightlist.append(int(line.split(',')[1])/10)
            weight = int(line.split(',')[1])/10
            print(str(weight))
            label['text'] = ""+str(weight)+"kg"
            if int(line.split(',')[1])>300:
                if max(weightlist[-15:-1])-min(weightlist[-15:])<2:
                    weight=sum(weightlist[-15:])/15
                    weight=round(weight,1)
                    print(weight)
                    exitnum=1
                    ser.write(b"9")
                else:
                    pass
            else:
                pass

        ws = wb.active
        enumber = 1
        while not ws.cell(enumber, 1).value is None:
            enumber += 1
        enumber -= 1
        ws.cell(enumber, 8).value = float(weight)
        wb.save('hikensya.xlsx')
        print(weight)

        label['text'] = str(weight)
        weight=label['text']
        print(weight)

        buttonEnter = tk.Button(frame_kg,text = '決定', command=click_btnEnter,font=("",50))
        buttonEnter.place(x=50, y=205, relwidth=0.36, relheight= 0.125)
        label2.destroy()
        label3 = tk.Label(frame_kg, text="決定を押してください",font=("",30),fg="red")
        label3.place(x=40, y=0)

    def click_skip():
        global BodyWeight
        BodyWeight = float(1000)

        label['text'] = ""+str(weight)+"kg"
        buttonEnter = tk.Button(frame_kg,text = '決定', command=click_btnEnter,font=("",50))
        buttonEnter.place(x=50, y=205, relwidth=0.36, relheight= 0.125)
        label2.destroy()
        label3 = tk.Label(frame_kg, text="決定を押してください",font=("",30),fg="red")
        label3.place(x=40, y=0)
        buttonSkip.destroy()

    frame_kg = tk.Frame(baseGround, width=500, height=600)
    frame_kg.pack()

    # ラベルやボタンの作成と配置

    label2 = tk.Label(frame_kg, text="台に乗ってください",font=("",30))
    label2.place(x=60, y=0)

    label = tk.Label(frame_kg, text='0', background='lightgray',font=("",50))
    label.place(x=50, y=50, width=400, height=105)

    buttonFinish = tk.Button(frame_kg,text = '戻る', command=click_btnFinish)
    buttonFinish.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

    buttonSkip = tk.Button(frame_kg,text = 'スキップ', command=click_skip)
    buttonSkip.place(x=359, y=300, relwidth=0.18, relheight= 0.125)

    thread2 = threading.Thread(target=BodyWeightSave)
    thread2.start()

def step5_scan():
    def click_btnFinish():
        frame_scan.destroy()
        step6_jushin()

    def click_btnEnter():
        thread1 = threading.Thread(target=start_scan)
        thread1.start()

    def start_scan():
        global file_cope
        ser.write(b"3") # シリアル通信でArduino に case3 のスキャン命令
        buttonEnter.destroy()
        # buttonFinish.destroy()
        label3 = tk.Label(frame_scan, text='スキャン中', background='lightgray',font=("",25))
        label3.place(x=50, y=90, width=400, height=105)
        exitnum=10
        # exitnum=11
        while exitnum==10:
            line2 = int(ser.readline())
            if line2 !=11:
                pass
            else:
                exitnum=11
        label.destroy()
        label2.destroy()
        label3.destroy()

        list_of_files = glob.glob('C:/Users/BRLAB/Pictures/*.jpg') # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime) # key=os.path.getctime は更新日時、
        karinumber = latest_file.removeprefix('C:/Users/BRLAB/Pictures\img')
        karinumber = karinumber.removesuffix('.jpg')

        print(latest_file)
        print(enumber-1)
        print(jnum)
        print(idnum)
        print(seibetsu)
        print(nenrei)
        print(shincho)


        rename_file='C:/Users/BRLAB/Desktop/photo/'+ str(enumber-1)+'_'+str(jnum)+'_'+str(idnum)+'_'+str(today)+'.jpg'
        file_cope = str(enumber-1)+'_'+str(jnum)+'_'+str(idnum)+'_'+str(today)
        os.rename(latest_file,rename_file) # os.rename(変更前ファイル、変更後ファイル)

        label4 = tk.Label(frame_scan, text="スキャンが完了しました\n終了を押してください",font=("",30))
        label4.place(x=50, y=0)

        label5 = tk.Label(frame_scan, text='スキャン完了', background='lightgray',font=("",25))
        label5.place(x=50, y=90, width=400, height=105)

        buttonFinish = tk.Button(frame_scan,text = '重心測定へ', command=click_btnFinish)
        buttonFinish.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

    def click_skip():
        buttonEnter.destroy()
        # buttonFinish.destroy()
        label3 = tk.Label(frame_scan, text='スキャン中', background='lightgray',font=("",25))
        label3.place(x=50, y=90, width=400, height=105)


        label4 = tk.Label(frame_scan, text="スキャンが完了しました\n終了を押してください",font=("",30))
        label4.place(x=50, y=0)

        label5 = tk.Label(frame_scan, text='スキャン完了', background='lightgray',font=("",25))
        label5.place(x=50, y=90, width=400, height=105)

        buttonFinish = tk.Button(frame_scan,text = '重心測定へ', command=click_btnFinish)
        buttonFinish.place(x=359, y=205, relwidth=0.18, relheight= 0.125)
        buttonSkip.destroy()
    frame_scan = tk.Frame(baseGround, width=500, height=600)
    frame_scan.pack()

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_scan, text="足裏をスキャンします\nスキャンを押してください",font=("",30))
    label2.place(x=50, y=0)

    label = tk.Label(frame_scan, text='スキャン準備完了', background='lightgray',font=("",25))
    label.place(x=50, y=90, width=400, height=105)

    buttonEnter = tk.Button(frame_scan,text = 'スキャン', command=click_btnEnter,font=("",30),fg='red')
    buttonEnter.place(x=50, y=205, relwidth=0.36, relheight= 0.125)

    buttonSkip = tk.Button(frame_scan,text = 'スキップ', command=click_skip)
    buttonSkip.place(x=359, y=300, relwidth=0.18, relheight= 0.125)
    # buttonFinish = tk.Button(frame_scan,text = '終了', command=click_btnFinish)
    # buttonFinish.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

def step6_jushin(): # 重心動揺測定！！ ##############################################################
    frame_jushin = tk.Frame(baseGround, width=500, height=600)
    frame_jushin.pack()
    def click_btnEnter():
        frame_jushin.destroy()
        step7_finish_sokutei()

    def timer():
      for i in range(5):
        label['text']=""+str(5-i)+"秒"
        time.sleep(0.5)
      thread4 = threading.Thread(target=jushin_sokutei)
      thread4.start()
      for i in range(35):
        label['text']=""+str(35-i)+"秒"
        time.sleep(0.2)
      label['text']="計測終了"
    #   ser.write(b"9")
      buttonEnter = tk.Button(frame_jushin,text = '決定', command=click_btnEnter,font=("",50))
      buttonEnter.place(x=50, y=250, relwidth=0.36, relheight= 0.125)
      buttonEnter['fg']="red"
      label4 = tk.Label(frame_jushin, text="重心計測が終了しました\n決定を押してください",font=("",30))
      label4.place(x=55, y=0)
      label4['bg']="gray94"

    def jushin_sokutei():
        global jushinlist
        # ser.write(b"4") # Arduino にケース４を命令、重心測定
        label2.destroy()
        label3 = tk.Label(frame_jushin, text="重心を計測しています\n終了まで",font=("",30))
        label3.place(x=50, y=0)
        label3['fg']="black"
        enumber=1

        ws = wb.active
        while not ws.cell(1, enumber).value is None:
            enumber =enumber + 1
        enumber=enumber-1

        jushinlist = [[0 for i in range(3000)] for j in range(2)]

        # for i in range(3000):
        #     line = str(ser.readline())
        #     print(line)
        #     try:
        #         jushinlist[0][i]=int(line.split(',')[1]) # xposition
        #         jushinlist[1][i]=int(line.split(',')[2]) # yposition
        #         # Arduino でweight1~4を使い重心の(pointX,pointY)を計算している。それをser.を使って読み込んでいる。
        #     except:
        #         jushinlist[0][i]=99999
        #         jushinlist[1][i]=99999

        # filename = 'C:/Users/BRLAB/Desktop/jushin/'+ str(file_cope) +'.csv'
        # with open(str(filename), 'a',newline="") as csv_file:
        #     fieldnames = ['time','xposition', 'yposition']
        #     writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        #     writer.writeheader()
        #     for i in range(3000):
        #         writer.writerow({'time': i,'xposition':  jushinlist[0][i], 'yposition': jushinlist[1][i]})


        # val=2

        # #ローパスフィルタ 今回使用
        # def lowpass(x, samplerate, fp, fs, gpass, gstop):
        #     fn = samplerate / 2                           #ナイキスト周波数
        #     wp = fp / fn                                  #ナイキスト周波数で通過域端周波数を正規化
        #     ws = fs / fn                                  #ナイキスト周波数で阻止域端周波数を正規化
        #     N, Wn = signal.buttord(wp, ws, gpass, gstop)  #オーダーとバターワースの正規化周波数を計算
        #     b, a = signal.butter(N, Wn, "low")            #フィルタ伝達関数の分子と分母を計算
        #     y = signal.filtfilt(b, a, x)                  #信号に対してフィルタをかける
        #     return y

        # # csvから列方向に順次フィルタ処理を行い保存する関数
        # def csv_filter(in_file, out_file, type):
        #     df = pd.read_csv(in_file, encoding='SHIFT-JIS')                  # ファイル読み込み
        #     dt = 1/80                                              # 時間刻み

        #     # データフレームを初期化
        #     df_filter = pd.DataFrame()
        #     df_filter[df.columns[0]] = df.T.iloc[0]

        #     # ローパスの設定-----------------------------------------------------------------------------
        #     # fp_lp = 15                                                       # 通過域端周波数[Hz]
        #     # fs_lp = 30                                                       # 阻止域端周波数[Hz]

        #     fp_lp = 3                                                      # 通過域端周波数[Hz]
        #     fs_lp = 5                                                       # 阻止域端周波数[Hz]

        #     gpass = 3                                                        # 通過域端最大損失[dB]
        #     gstop = 40                                                       # 阻止域端最小損失[dB]

        #     # 列方向に順次フィルタ処理をするコード
        #     for i in range(len(df.T)-1):
        #         data = df.T.iloc[i+1]                       # フィルタ処理するデータ列を抽出

        #         # フィルタ処理の種類を文字列で読み取って適切な関数を選択する
        #         if type == 'lp':
        #             # ローパスフィルタを実行
        #             # print('wave=', i, ':Lowpass.')
        #             data = lowpass(x=data, samplerate=1 / dt,
        #                         fp=fp_lp, fs=fs_lp,
        #                         gpass=gpass, gstop=gstop)
        #         else:
        #             # 文字列が当てはまらない時はパス（動作テストでフィルタかけたくない時はNoneとか書いて実行するとよい）
        #             pass

        #         data = pd.Series(data)                                       # dataをPandasシリーズデータへ変換
        #         df_filter[df.columns[i + 1] + '_filter'] = data              # 保存用にデータフレームへdataを追加

        #     df_filter.to_csv(out_file)                                       # フィルタ処理の結果をcsvに保存

        #     return df, df_filter

        # # p=filename
        # # p=str(p.split('.')[0])
        # # print(p)
        # df, df_filter = csv_filter(in_file='C:/Users/BRLAB/Desktop/jushin/'+str(file_cope)+'.csv', out_file='C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv', type='lp')
        # # ws.cell(val, 1).value = p
        # #print(ws.cell(1, 2).value) (1,b)
        # with open('C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv', 'r') as csvfile:
        # # with open('101_1_0_20211204.csv', 'r') as csvfile:
        #     # print(p)
        #     reader = csv.reader(csvfile)
        #     csv_data = list(reader)
        #     # sum=((int(csv_data[1][1])-int(csv_data[2][1]))**2+(int(csv_data[1][2])-int(csv_data[2][2]))**2)**0.5
        #     # print(sum)
        #     # print(csv_data[1][1])
        #     # print(csv_data[1][2])
        #     sum_all=0
        #     sum_all=float(sum_all)
        #     sum_x=0
        #     sum_x=float(sum_x)
        #     sum_y=0
        #     sum_y=float(sum_y)
        #     df = pd.read_csv('C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv',index_col=0)
        #     max_x=df.max()['xposition_filter']
        #     min_x=df.min()['xposition_filter']
        #     maxd_x=max_x-min_x
        #     max_y=df.max()['yposition_filter']
        #     min_y=df.min()['yposition_filter']
        #     maxd_y=max_y-min_y
        #     for i in range(2400):
        #         sum_all=sum_all+((float(csv_data[i+1][2])-float(csv_data[i+2][2]))**2+(float(csv_data[i+1][3])-float(csv_data[i+2][3]))**2)**0.5
        #         sum_x=sum_x+abs(float(csv_data[i+1][2])-float(csv_data[i+2][2]))
        #         sum_y=sum_y+abs(float(csv_data[i+1][3])-float(csv_data[i+2][3]))
        #     # print(sum_all)
        # # ws.cell(val, 6).value = str(sum_all)
        # # ws.cell(val, 7).value = str(sum_x)
        # # ws.cell(val, 8).value = str(sum_y)
        # # ws.cell(val, 9).value = str(maxd_x)
        # # ws.cell(val, 10).value = str(maxd_y)
        # #print(ws.cell(1, 2).value) (1,b)
        # # wb.save('jushin_kekka.xlsx')
        # val=val+1

        # fig = plt.figure(figsize=[15,15])

        # input_csv = pd.read_csv('C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv')
        # first_column_data = input_csv[input_csv.keys()[2]]
        # second_column_data = input_csv[input_csv.keys()[3]]

        # plt.xlabel(input_csv.keys()[2],fontsize=20)
        # plt.ylabel(input_csv.keys()[3],fontsize=20)

        # plt.plot(first_column_data, second_column_data, linestyle='solid', marker='')
        # plt.tick_params(labelsize=20)
        # #plt.show()
        # fig.savefig('C:/Users/BRLAB/Desktop/jushin/csv_filter_photo/'+str(file_cope)+"_filter.png")


        # print("総軌跡長＝",sum_all)
        # print("x最大変位=",maxd_x)
        # print("y最大変位＝",maxd_y)

        # xy_area = maxd_x*maxd_y
        # print("矩形面積＝",xy_area)

        # ws = wb.active
        # enumber = 1
        # while not ws.cell(enumber, 1).value is None:
        #     enumber += 1
        # enumber -= 1
        # ws.cell(enumber, 9).value = float(sum_all)
        # ws.cell(enumber, 10).value = float(maxd_x)
        # ws.cell(enumber, 11).value = float(maxd_y)
        # ws.cell(enumber, 12).value = float(xy_area)


        # wb.save('/Users/BRLAB/hikensya.xlsx')

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_jushin, text="重心を35秒間測定します\n直立を維持してください\n開始まで",font=("",30))
    label2.place(x=30, y=0)
    label2['fg']="black"

    label = tk.Label(frame_jushin, text='0', background='lightgray',font=("",50))
    label.place(x=50, y=130, width=400, height=105)


    thread3 = threading.Thread(target=timer)
    thread3.start()

# 重心動揺測定終了後、「はじめに戻る」もしくは「画像処理/解析」のボタン表示 ###########################################################3
def step7_finish_sokutei(): # 測定終了画面
    frame_fini = tk.Frame(baseGround, width=500, height=600) # 測定終了画面、フレーム、
    frame_fini.pack() # 配置、特に指定なし
    def click_btnStart1():   # buttonStart1 が押されたときのイベント
        global jnum
        jnum=1
        frame_fini.destroy()  # 測定終了画面を消す
        step1_start()        # start_jouken を実行

    def click_btnStart2(): # buttonStart2 が押されたときのイベント
        frame_fini.destroy()  # 測定終了画面を消す
        kaiseki_gui()         # kaiseki_guiを実行

    def click_btnStart3(): # buttonStart3 が押されたときのイベント
        frame_fini.destroy()  # 測定終了画面を消す
        shindan1()            # shindan1 を実行
    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_fini, text="計測終了です。",font=("",30))
    label2.place(x=120, y=100)
    label2['fg']="black"

    buttonStart1 = tk.Button(frame_fini,text = 'はじめに戻る', command=click_btnStart1,font=("",20))
    buttonStart1.place(x=50, y=195, relwidth=0.75, relheight= 0.125)

    buttonStart2 = tk.Button(frame_fini,text = '画像処理/解析', command=click_btnStart2,font=("",20))
    buttonStart2.place(x=50, y=295, relwidth=0.75, relheight= 0.125)

    # buttonStart3 = tk.Button(frame_fini,text = '診断', command=click_btnStart3,font=("",20))
    # buttonStart3.place(x=50, y=395, relwidth=0.75, relheight= 0.125)

# メインウィンドウを作成
baseGround = tk.Tk()
# ウィンドウのサイズを設定
# baseGround.geometry('500x600')
baseGround.state('zoomed')
# 画面タイトル
baseGround.title('足裏測定')
step1_start()

baseGround.mainloop()