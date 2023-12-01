import tkinter as tk
import openpyxl as px  # excel 書き込み用、pip install openpyxl 必要
import datetime
import time
import threading

examinee_xlsx = 'C:/Users/zutom/.vscode/PythonFile/FootScan/Practice/examinee.xlsx'
wb = px.load_workbook(examinee_xlsx) # 被験者情報の保存先のエクセルファイル
ws = wb.active

def step1_start():
    frame_start = tk.Frame(baseGround, width=500, height=700)
    frame_start.pack()
    def click_btnStart1(): # 下のスタートボタン1を押した後のイベント
        global jnum
        jnum=1
        frame_start.destroy()
        input_1_ID()
    def click_btnStart2(): # 下のスタートボタン2を押した後のイベント
        global jnum
        jnum=2
        frame_start.destroy()
        input_1_ID()
    def click_btnStart3(): # 下のスタートボタン3を押した後のイベント
        global jnum
        jnum=3
        frame_start.destroy()
        input_1_ID()
    def click_btnStart4(): # 下のスタートボタン4を押した後のイベント
        global jnum
        jnum=4
        frame_start.destroy()
        input_1_ID()
    def file_dialog(): # 下のスタートボタン5を押した後のイベント
        frame_start.destroy()
        kaiseki_gui()

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_start, text="まだ台に乗らないでください",font=("",30))
    label2.place(x=20, y=10)
    label2['fg']="red"

    buttonStart1 = tk.Button(frame_start,text = '脱いですぐSTART', command=click_btnStart1,font=("",20))
    buttonStart1.place(x=50, y=145, relwidth=0.75, relheight= 0.125)

    buttonStart2 = tk.Button(frame_start,text = '脱いでしばらくしてSTART', command=click_btnStart2,font=("",20))
    buttonStart2.place(x=50, y=245, relwidth=0.75, relheight= 0.125)

    buttonStart3 = tk.Button(frame_start,text = '+20kgでSTART', command=click_btnStart3,font=("",20))
    buttonStart3.place(x=50, y=345, relwidth=0.75, relheight= 0.125)

    buttonStart4 = tk.Button(frame_start,text = 'ウエットティッシュでSTART', command=click_btnStart4,font=("",20))
    buttonStart4.place(x=50, y=445, relwidth=0.75, relheight= 0.125)

    buttonStart5 = tk.Button(frame_start,text = '画像処理/解析', command=file_dialog,font=("",20))
    buttonStart5.place(x=50, y=545, relwidth=0.75, relheight= 0.125)

# ここから被験者番号入力 #####################################################################
def input_1_ID(): # 被検者番号入力用GUIを一つの関数に
  global frame_id # frame_id(被験者番号の入力値) を他と共有できるように、、global!
  frame_id=0
  # 被検者番号の数字入力用ボタンを押したあとの動作を定義、click_btn1…はあとで使う
  def click_btn1(): # 159行目のbutton1("1"のボタン)を押されたときのイベント
      if label['text'] == '0': # label は156行目で定義した文字表示用ラベル、これを「1」に、(47行目)もしくは「元の数字+1」にする
          label['text'] = '1'
          buttonEnter['fg'] = "red" # buttonEnterは192行目で定義した「決定」のボタン、赤色に染めるエフェクト?
          label2['fg']="black" # label2は153行目に定義した「被検者番号を入力してください」のラベル
      else:
          label['text'] = label['text'] +'1'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn2():
      if label['text'] == '0':
          label['text'] = '2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn3():
      if label['text'] == '0':
          label['text'] = '3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn4():
      if label['text'] == '0':
          label['text'] = '4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn5():
      if label['text'] == '0':
          label['text'] = '5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn6():
      if label['text'] == '0':
          label['text'] = '6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn7():
      if label['text'] == '0':
          label['text'] = '7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn8():
      if label['text'] == '0':
          label['text'] = '8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn9():
      if label['text'] == '0':
          label['text'] = '9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn0():
      if label['text'] == '0':
          label['text'] = '0'
      else:
          label['text'] = label['text'] +'0'
          buttonEnter['fg'] = "red"
          label2['fg']="black"

  def click_btnAc():
      label['text'] = '0'
      buttonEnter['fg'] = "black"
      label2['fg']="black"

  def click_btnFinish():
    frame_id.destroy()
    step1_start()

  def click_btnEnter():
    global idnum
    idnum=0
    idnum=int(label['text'])
    print(idnum)
    frame_id.destroy()
    input_2_gender()

  baseGround.title('足裏測定')

  # フレームサイズ定義
  frame_id = tk.Frame(baseGround, width=500, height=600)
  frame_id.pack() # 配置をpackに
  # ラベルやボタンの作成と配置
  label2 = tk.Label(frame_id, text="被検者番号を入力してください",font=("",30))
  label2.place(x=25, y=0)

  label = tk.Label(frame_id, text='0', background='lightgray',font=("",50))
  label.place(x=50, y=50, width=400, height=105)

  button1 = tk.Button(frame_id,text = '1', command=click_btn1,font=("",50)) # 上で定義したclick_btn1…を装備
  button1.place(x=50, y=205, relwidth=0.18, relheight= 0.125)

  button2 = tk.Button(frame_id,text = '2', command=click_btn2,font=("",50)) # 上で定義したclick_btn2…を装備
  button2.place(x=153, y=205, relwidth=0.18, relheight= 0.125)

  button3 = tk.Button(frame_id,text = '3', command=click_btn3,font=("",50)) #
  button3.place(x=256, y=205, relwidth=0.18, relheight= 0.125)

  button4 = tk.Button(frame_id,text = '4', command=click_btn4,font=("",50))
  button4.place(x=50, y=295, relwidth=0.18, relheight= 0.125)

  button5 = tk.Button(frame_id,text = '5', command=click_btn5,font=("",50))
  button5.place(x=153, y=295, relwidth=0.18, relheight= 0.125)

  button6 = tk.Button(frame_id,text = '6', command=click_btn6,font=("",50))
  button6.place(x=256, y=295, relwidth=0.18, relheight= 0.125)

  buttonFinish = tk.Button(frame_id,text = '終了', command=click_btnFinish,font=("",15))
  buttonFinish.place(x=359, y=295, relwidth=0.18, relheight= 0.125)

  button7 = tk.Button(frame_id,text = '7', command=click_btn7,font=("",50))
  button7.place(x=50, y=385, relwidth=0.18, relheight= 0.125)

  button8 = tk.Button(frame_id,text = '8', command=click_btn8,font=("",50))
  button8.place(x=153, y=385, relwidth=0.18, relheight= 0.125)

  button9 = tk.Button(frame_id,text = '9', command=click_btn9,font=("",50))
  button9.place(x=256, y=385, relwidth=0.18, relheight= 0.125)

  button0 = tk.Button(frame_id,text = '0', command=click_btn0,font=("",50))
  button0.place(x=153, y=475, relwidth=0.18, relheight= 0.125)

  buttonEnter = tk.Button(frame_id,text = '決定', command=click_btnEnter,font=("",50))
  buttonEnter.place(x=256, y=475, relwidth=0.38, relheight= 0.125)

  buttonAc = tk.Button(frame_id,text = 'AC', command=click_btnAc,font=("",50))
  buttonAc.place(x=50, y=475, relwidth=0.18, relheight= 0.125)

# ここから年齢入力 #########################################################################
def input_3_age():
  global frame
  # 年齢の数字入力用ボタンを押したあとの動作を定義、click_btn1…はあとで使う
  def click_btn1():
      if label['text'] == '0':
          label['text'] = '1'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'1'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn2():
      if label['text'] == '0':
          label['text'] = '2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn3():
      if label['text'] == '0':
          label['text'] = '3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn4():
      if label['text'] == '0':
          label['text'] = '4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn5():
      if label['text'] == '0':
          label['text'] = '5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn6():
      if label['text'] == '0':
          label['text'] = '6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn7():
      if label['text'] == '0':
          label['text'] = '7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn8():
      if label['text'] == '0':
          label['text'] = '8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn9():
      if label['text'] == '0':
          label['text'] = '9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn0():
      if label['text'] == '0':
          label['text'] = '0'
      else:
          label['text'] = label['text'] +'0'
          buttonEnter['fg'] = "red"
          label2['fg']="black"

  def click_btnAc():
      label['text'] = '0'
      buttonEnter['fg'] = "black"
      label2['fg']="black"

  def click_btnReturn():
    frame.destroy()
    input_2_gender()

  def click_btnFinish():
    frame.destroy()
    step1_start()

  def click_btnEnter():
    global nenrei
    nenrei=0

    if( int(label['text']) > 15 and int(label['text']) < 120 ):
        nenrei=int(label['text'])
        frame.destroy()
        input_4_height()
    else:
        label2['fg']="red"
        buttonEnter['fg'] = "black"


  baseGround.title('足裏測定')

  frame = tk.Frame(baseGround, width=500, height=600)
  frame.pack()
  # ラベルやボタンの作成と配置
  label2 = tk.Label(frame, text="年齢を入力してください",font=("",30))
  label2.place(x=40, y=0)

  label = tk.Label(frame, text='0', background='lightgray',font=("",50))
  label.place(x=50, y=50, width=400, height=105)

  button1 = tk.Button(frame,text = '1', command=click_btn1,font=("",50))
  button1.place(x=50, y=205, relwidth=0.18, relheight= 0.125)

  button2 = tk.Button(frame,text = '2', command=click_btn2,font=("",50))
  button2.place(x=153, y=205, relwidth=0.18, relheight= 0.125)

  button3 = tk.Button(frame,text = '3', command=click_btn3,font=("",50))
  button3.place(x=256, y=205, relwidth=0.18, relheight= 0.125)

  buttonReturn = tk.Button(frame,text = '性別入力\nに戻る', command=click_btnReturn,font=("",15))
  buttonReturn.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

  button4 = tk.Button(frame,text = '4', command=click_btn4,font=("",50))
  button4.place(x=50, y=295, relwidth=0.18, relheight= 0.125)

  button5 = tk.Button(frame,text = '5', command=click_btn5,font=("",50))
  button5.place(x=153, y=295, relwidth=0.18, relheight= 0.125)

  button6 = tk.Button(frame,text = '6', command=click_btn6,font=("",50))
  button6.place(x=256, y=295, relwidth=0.18, relheight= 0.125)

  buttonFinish = tk.Button(frame,text = '終了', command=click_btnFinish,font=("",15))
  buttonFinish.place(x=359, y=295, relwidth=0.18, relheight= 0.125)

  button7 = tk.Button(frame,text = '7', command=click_btn7,font=("",50))
  button7.place(x=50, y=385, relwidth=0.18, relheight= 0.125)

  button8 = tk.Button(frame,text = '8', command=click_btn8,font=("",50))
  button8.place(x=153, y=385, relwidth=0.18, relheight= 0.125)

  button9 = tk.Button(frame,text = '9', command=click_btn9,font=("",50))
  button9.place(x=256, y=385, relwidth=0.18, relheight= 0.125)

  button0 = tk.Button(frame,text = '0', command=click_btn0,font=("",50))
  button0.place(x=153, y=475, relwidth=0.18, relheight= 0.125)

  buttonEnter = tk.Button(frame,text = '決定', command=click_btnEnter,font=("",50))
  buttonEnter.place(x=256, y=475, relwidth=0.38, relheight= 0.125)

  buttonAc = tk.Button(frame,text = 'AC', command=click_btnAc,font=("",50))
  buttonAc.place(x=50, y=475, relwidth=0.18, relheight= 0.125)

# ここから身長入力 ######################################################################
def input_4_height():
  global frame_app
  global jushinlist
  # 身長の数字入力用ボタンを押したあとの動作を定義、click_btn1…はあとで使う
  def click_btn1():
      if label['text'] == '0':
          label['text'] = '1'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'1'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn2():
      if label['text'] == '0':
          label['text'] = '2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'2'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn3():
      if label['text'] == '0':
          label['text'] = '3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'3'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn4():
      if label['text'] == '0':
          label['text'] = '4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'4'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn5():
      if label['text'] == '0':
          label['text'] = '5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'5'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn6():
      if label['text'] == '0':
          label['text'] = '6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'6'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn7():
      if label['text'] == '0':
          label['text'] = '7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'7'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn8():
      if label['text'] == '0':
          label['text'] = '8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'8'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn9():
      if label['text'] == '0':
          label['text'] = '9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
      else:
          label['text'] = label['text'] +'9'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btn0():
      if label['text'] == '0':
          label['text'] = '0'
      else:
          label['text'] = label['text'] +'0'
          buttonEnter['fg'] = "red"
          label2['fg']="black"
  def click_btnAc():
      label['text'] = '0'
      buttonEnter['fg'] = "black"
      label2['fg']="black"
  def click_btnReturn(): # 戻るボタンで年齢入力に戻る
    frame_app.destroy()
    input_3_age()
  def click_btnFinish(): # 終了ボタンでstart_jouken() 一番初めに戻る
    frame_app.destroy()
    step1_start()
  def click_btnEnter(): # 決定ボタンで、被験者情報をエクセルに書き込み、体重測定開始
    global today
    global shincho
    global enumber
    enumber=1

    # if( int(label['text']) > 100 and int(label['text']) < 250 ):
    #     shincho=int(label['text'])

    #     dt_now = datetime.datetime.now() # 今日の日付
    #     yearnum=dt_now.year
    #     monthnum=dt_now.month
    #     monthnum = '{0:0>2}'.format(monthnum)
    #     daynum=dt_now.day
    #     daynum = '{0:0>2}'.format(daynum)
    #     today=str(yearnum)+str(monthnum)+str(daynum)

    #     while not ws.cell(enumber, 1).value is None: # hikensya.xlsx でデータの無い行まで１からプラス１していく（必然と最後尾に付く）
    #         enumber += 1
    #     ws.cell(enumber, 1).value = int(enumber)-1 # ここから下、被験者情報をエクセルに保存、cell(enumber, 1)はenumber行 1 列目のセルを指定
    #     ws.cell(enumber, 2).value = int(today)
    #     ws.cell(enumber, 3).value = int(jnum)
    #     ws.cell(enumber, 4).value = int(idnum)
    #     ws.cell(enumber, 5).value = int(seibetsu)
    #     ws.cell(enumber, 6).value = int(nenrei)
    #     ws.cell(enumber, 7).value = int(shincho)
    #     #ws.cell(enumber, 8).value = float(weight)
    #     wb.save('hikensya.xlsx')


    frame_app.destroy()
    step3_BodyWeightCali() # 体重測定に移行するようコール
    # start()
    # baseGround.destroy()
# else:
#     label2['fg']="red"
#     buttonEnter['fg'] = "black"

  baseGround.title('足裏測定')

  frame_app = tk.Frame(baseGround, width=500, height=600)
  frame_app.pack()
  # ラベルやボタンの作成と配置

  label2 = tk.Label(frame_app, text="身長(cm)を入力してください",font=("",30))
  label2.place(x=35, y=0)

  label = tk.Label(frame_app, text='0', background='lightgray',font=("",50))
  label.place(x=50, y=50, width=400, height=105)

  button1 = tk.Button(frame_app,text = '1', command=click_btn1,font=("",50))
  button1.place(x=50, y=205, relwidth=0.18, relheight= 0.125)

  button2 = tk.Button(frame_app,text = '2', command=click_btn2,font=("",50))
  button2.place(x=153, y=205, relwidth=0.18, relheight= 0.125)

  button3 = tk.Button(frame_app,text = '3', command=click_btn3,font=("",50))
  button3.place(x=256, y=205, relwidth=0.18, relheight= 0.125)

  buttonReturn = tk.Button(frame_app,text = '年齢入力\nに戻る', command=click_btnReturn,font=("",15))
  buttonReturn.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

  button4 = tk.Button(frame_app,text = '4', command=click_btn4,font=("",50))
  button4.place(x=50, y=295, relwidth=0.18, relheight= 0.125)

  button5 = tk.Button(frame_app,text = '5', command=click_btn5,font=("",50))
  button5.place(x=153, y=295, relwidth=0.18, relheight= 0.125)

  button6 = tk.Button(frame_app,text = '6', command=click_btn6,font=("",50))
  button6.place(x=256, y=295, relwidth=0.18, relheight= 0.125)

  buttonFinish = tk.Button(frame_app,text = '終了', command=click_btnFinish,font=("",15))
  buttonFinish.place(x=359, y=295, relwidth=0.18, relheight= 0.125)

  button7 = tk.Button(frame_app,text = '7', command=click_btn7,font=("",50))
  button7.place(x=50, y=385, relwidth=0.18, relheight= 0.125)

  button8 = tk.Button(frame_app,text = '8', command=click_btn8,font=("",50))
  button8.place(x=153, y=385, relwidth=0.18, relheight= 0.125)

  button9 = tk.Button(frame_app,text = '9', command=click_btn9,font=("",50))
  button9.place(x=256, y=385, relwidth=0.18, relheight= 0.125)

  button0 = tk.Button(frame_app,text = '0', command=click_btn0,font=("",50))
  button0.place(x=153, y=475, relwidth=0.18, relheight= 0.125)

  buttonEnter = tk.Button(frame_app,text = '決定', command=click_btnEnter,font=("",50))
  buttonEnter.place(x=256, y=475, relwidth=0.38, relheight= 0.125)

  buttonAc = tk.Button(frame_app,text = 'AC', command=click_btnAc,font=("",50))
  buttonAc.place(x=50, y=475, relwidth=0.18, relheight= 0.125)

# ここから性別入力 ##############################################################
def input_2_gender():
    def click_btn1():
        if label['text'] == '女':
            label['text'] = '男'
            buttonEnter['fg'] = "red"
            label2['fg']="black"

        else:
            label['text'] = '男'
            buttonEnter['fg'] = "red"
            label2['fg']="black"

    def click_btn2():
        if label['text'] == '男':
            label['text'] = '女'
            buttonEnter['fg'] = "red"
            label2['fg']="black"

        else:
            label['text'] = '女'
            buttonEnter['fg'] = "red"
            label2['fg']="black"

    def click_btnFinish():
        frame_ori.destroy()
        step1_start()

    def click_btnEnter():
        global seibetsu
        seibetsu=0
        if label['text'] == '男':
            seibetsu=1
            frame_ori.destroy()
            input_3_age()
        elif label['text'] == '女':
            seibetsu=2
            frame_ori.destroy()
            input_3_age()
        else:
            label2['fg']="red"

    frame_ori = tk.Frame(baseGround, width=500, height=600)
    frame_ori.pack()
    # ラベルやボタンの作成と配置

    label2 = tk.Label(frame_ori, text="性別を入力してください",font=("",30))
    label2.place(x=40, y=0)

    label = tk.Label(frame_ori, text='', background='lightgray',font=("",50))
    label.place(x=50, y=50, width=400, height=105)

    button1 = tk.Button(frame_ori,text = '男', command=click_btn1,font=("",50))
    button1.place(x=50, y=205, relwidth=0.18, relheight= 0.125)

    button2 = tk.Button(frame_ori,text = '女', command=click_btn2,font=("",50))
    button2.place(x=153, y=205, relwidth=0.18, relheight= 0.125)

    buttonEnter = tk.Button(frame_ori,text = '決定', command=click_btnEnter,font=("",50))
    buttonEnter.place(x=256, y=205, relwidth=0.36, relheight= 0.125)

    buttonFinish = tk.Button(frame_ori,text = '終了', command=click_btnFinish)
    buttonFinish.place(x=359, y=295, relwidth=0.18, relheight= 0.125)

def step3_BodyWeightCali():

    frame_cali = tk.Frame(baseGround, width=800, height=600)
    frame_cali.pack()

    def timer():
        for i in range(8):
            label['text']=""+str(7-i)+"秒"
            time.sleep(0.5)
        frame_cali.destroy()
        step4_BodyWeight()

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_cali, text="まだ台に乗らないでください\n処理中です\n残り時間",font=("",30))
    label2.place(x=30, y=0)
    label2['fg']="red"

    label = tk.Label(frame_cali, text='0', background='lightgray',font=("",50))
    label.place(x=50, y=130, width=400, height=105)
    thread3 = threading.Thread(target=timer)
    thread3.start()

def step4_BodyWeight():
    def click_btnFinish():
        # ser.write(b"9") # case1 内の体重測定コール（'9'）
        # ser.close()
        frame_kg.destroy()
        step3_BodyWeightCali()

    def click_btnEnter(): # 決定ボタンが押された後のイベント（スキャンへ）
        frame_kg.destroy()
        step5_scan()

    def BodyWeightSave():
        global weight
        weight=0
        weightlist=[0,0]
        exitnum=0
        while exitnum==0:
            line = str(ser.readline()) # readlineはファイルから1行だけ読み出す。readlinesはファイルの内容を全て読み出し、1行ごとのリストにします。
            weightlist.append(int(line.split(',')[1])/10)
            weight = int(line.split(',')[1])/10
            print(str(weight))
            label['text'] = ""+str(weight)+"kg"
            if int(line.split(',')[1])>300:
                if max(weightlist[-15:-1])-min(weightlist[-15:])<2:
                    weight=sum(weightlist[-15:])/15
                    weight=round(weight,1)
                    print(weight)
                    exitnum=1
                    ser.write(b"9")
                else:
                    pass
            else:
                pass

        ws = wb.active
        enumber = 1
        while not ws.cell(enumber, 1).value is None:
            enumber += 1
        enumber -= 1
        ws.cell(enumber, 8).value = float(weight)
        wb.save('hikensya.xlsx')
        print(weight)

        label['text'] = str(weight)
        weight=label['text']
        print(weight)

        buttonEnter = tk.Button(frame_kg,text = '決定', command=click_btnEnter,font=("",50))
        buttonEnter.place(x=50, y=205, relwidth=0.36, relheight= 0.125)
        label2.destroy()
        label3 = tk.Label(frame_kg, text="決定を押してください",font=("",30),fg="red")
        label3.place(x=40, y=0)

    def click_skip():
        global BodyWeight
        BodyWeight = float(1000)

        label['text'] = ""+str(weight)+"kg"
        buttonEnter = tk.Button(frame_kg,text = '決定', command=click_btnEnter,font=("",50))
        buttonEnter.place(x=50, y=205, relwidth=0.36, relheight= 0.125)
        label2.destroy()
        label3 = tk.Label(frame_kg, text="決定を押してください",font=("",30),fg="red")
        label3.place(x=40, y=0)
        buttonSkip.destroy()

    frame_kg = tk.Frame(baseGround, width=500, height=600)
    frame_kg.pack()

    # ラベルやボタンの作成と配置

    label2 = tk.Label(frame_kg, text="台に乗ってください",font=("",30))
    label2.place(x=60, y=0)

    label = tk.Label(frame_kg, text='0', background='lightgray',font=("",50))
    label.place(x=50, y=50, width=400, height=105)

    buttonFinish = tk.Button(frame_kg,text = '戻る', command=click_btnFinish)
    buttonFinish.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

    buttonSkip = tk.Button(frame_kg,text = 'スキップ', command=click_skip)
    buttonSkip.place(x=359, y=300, relwidth=0.18, relheight= 0.125)

    thread2 = threading.Thread(target=BodyWeightSave)
    thread2.start()

def step5_scan():
    def click_btnFinish():
        frame_scan.destroy()
        step6_jushin()

    def click_btnEnter():
        thread1 = threading.Thread(target=start_scan)
        thread1.start()

    def start_scan():
        global file_cope
        ser.write(b"3") # シリアル通信でArduino に case3 のスキャン命令
        buttonEnter.destroy()
        # buttonFinish.destroy()
        label3 = tk.Label(frame_scan, text='スキャン中', background='lightgray',font=("",25))
        label3.place(x=50, y=90, width=400, height=105)
        exitnum=10
        # exitnum=11
        while exitnum==10:
            line2 = int(ser.readline())
            if line2 !=11:
                pass
            else:
                exitnum=11
        label.destroy()
        label2.destroy()
        label3.destroy()

        list_of_files = glob.glob('C:/Users/BRLAB/Pictures/*.jpg') # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime) # key=os.path.getctime は更新日時、
        karinumber = latest_file.removeprefix('C:/Users/BRLAB/Pictures\img')
        karinumber = karinumber.removesuffix('.jpg')

        print(latest_file)
        print(enumber-1)
        print(jnum)
        print(idnum)
        print(seibetsu)
        print(nenrei)
        print(shincho)


        rename_file='C:/Users/BRLAB/Desktop/photo/'+ str(enumber-1)+'_'+str(jnum)+'_'+str(idnum)+'_'+str(today)+'.jpg'
        file_cope = str(enumber-1)+'_'+str(jnum)+'_'+str(idnum)+'_'+str(today)
        os.rename(latest_file,rename_file) # os.rename(変更前ファイル、変更後ファイル)

        label4 = tk.Label(frame_scan, text="スキャンが完了しました\n終了を押してください",font=("",30))
        label4.place(x=50, y=0)

        label5 = tk.Label(frame_scan, text='スキャン完了', background='lightgray',font=("",25))
        label5.place(x=50, y=90, width=400, height=105)

        buttonFinish = tk.Button(frame_scan,text = '重心測定へ', command=click_btnFinish)
        buttonFinish.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

    def click_skip():
        buttonEnter.destroy()
        # buttonFinish.destroy()
        label3 = tk.Label(frame_scan, text='スキャン中', background='lightgray',font=("",25))
        label3.place(x=50, y=90, width=400, height=105)


        label4 = tk.Label(frame_scan, text="スキャンが完了しました\n終了を押してください",font=("",30))
        label4.place(x=50, y=0)

        label5 = tk.Label(frame_scan, text='スキャン完了', background='lightgray',font=("",25))
        label5.place(x=50, y=90, width=400, height=105)

        buttonFinish = tk.Button(frame_scan,text = '重心測定へ', command=click_btnFinish)
        buttonFinish.place(x=359, y=205, relwidth=0.18, relheight= 0.125)
        buttonSkip.destroy()
    frame_scan = tk.Frame(baseGround, width=500, height=600)
    frame_scan.pack()

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_scan, text="足裏をスキャンします\nスキャンを押してください",font=("",30))
    label2.place(x=50, y=0)

    label = tk.Label(frame_scan, text='スキャン準備完了', background='lightgray',font=("",25))
    label.place(x=50, y=90, width=400, height=105)

    buttonEnter = tk.Button(frame_scan,text = 'スキャン', command=click_btnEnter,font=("",30),fg='red')
    buttonEnter.place(x=50, y=205, relwidth=0.36, relheight= 0.125)

    buttonSkip = tk.Button(frame_scan,text = 'スキップ', command=click_skip)
    buttonSkip.place(x=359, y=300, relwidth=0.18, relheight= 0.125)
    # buttonFinish = tk.Button(frame_scan,text = '終了', command=click_btnFinish)
    # buttonFinish.place(x=359, y=205, relwidth=0.18, relheight= 0.125)

def step6_jushin(): # 重心動揺測定！！ ##############################################################
    frame_jushin = tk.Frame(baseGround, width=500, height=600)
    frame_jushin.pack()
    def click_btnEnter():
        frame_jushin.destroy()
        step7_finish_sokutei()

    def timer():
      for i in range(5):
        label['text']=""+str(5-i)+"秒"
        time.sleep(0.5)
      thread4 = threading.Thread(target=jushin_sokutei)
      thread4.start()
      for i in range(35):
        label['text']=""+str(35-i)+"秒"
        time.sleep(0.2)
      label['text']="計測終了"
    #   ser.write(b"9")
      buttonEnter = tk.Button(frame_jushin,text = '決定', command=click_btnEnter,font=("",50))
      buttonEnter.place(x=50, y=250, relwidth=0.36, relheight= 0.125)
      buttonEnter['fg']="red"
      label4 = tk.Label(frame_jushin, text="重心計測が終了しました\n決定を押してください",font=("",30))
      label4.place(x=55, y=0)
      label4['bg']="gray94"

    def jushin_sokutei():
        global jushinlist
        # ser.write(b"4") # Arduino にケース４を命令、重心測定
        label2.destroy()
        label3 = tk.Label(frame_jushin, text="重心を計測しています\n終了まで",font=("",30))
        label3.place(x=50, y=0)
        label3['fg']="black"
        enumber=1

        ws = wb.active
        while not ws.cell(1, enumber).value is None:
            enumber =enumber + 1
        enumber=enumber-1

        jushinlist = [[0 for i in range(3000)] for j in range(2)]

        # for i in range(3000):
        #     line = str(ser.readline())
        #     print(line)
        #     try:
        #         jushinlist[0][i]=int(line.split(',')[1]) # xposition
        #         jushinlist[1][i]=int(line.split(',')[2]) # yposition
        #         # Arduino でweight1~4を使い重心の(pointX,pointY)を計算している。それをser.を使って読み込んでいる。
        #     except:
        #         jushinlist[0][i]=99999
        #         jushinlist[1][i]=99999

        # filename = 'C:/Users/BRLAB/Desktop/jushin/'+ str(file_cope) +'.csv'
        # with open(str(filename), 'a',newline="") as csv_file:
        #     fieldnames = ['time','xposition', 'yposition']
        #     writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        #     writer.writeheader()
        #     for i in range(3000):
        #         writer.writerow({'time': i,'xposition':  jushinlist[0][i], 'yposition': jushinlist[1][i]})


        # val=2

        # #ローパスフィルタ 今回使用
        # def lowpass(x, samplerate, fp, fs, gpass, gstop):
        #     fn = samplerate / 2                           #ナイキスト周波数
        #     wp = fp / fn                                  #ナイキスト周波数で通過域端周波数を正規化
        #     ws = fs / fn                                  #ナイキスト周波数で阻止域端周波数を正規化
        #     N, Wn = signal.buttord(wp, ws, gpass, gstop)  #オーダーとバターワースの正規化周波数を計算
        #     b, a = signal.butter(N, Wn, "low")            #フィルタ伝達関数の分子と分母を計算
        #     y = signal.filtfilt(b, a, x)                  #信号に対してフィルタをかける
        #     return y

        # # csvから列方向に順次フィルタ処理を行い保存する関数
        # def csv_filter(in_file, out_file, type):
        #     df = pd.read_csv(in_file, encoding='SHIFT-JIS')                  # ファイル読み込み
        #     dt = 1/80                                              # 時間刻み

        #     # データフレームを初期化
        #     df_filter = pd.DataFrame()
        #     df_filter[df.columns[0]] = df.T.iloc[0]

        #     # ローパスの設定-----------------------------------------------------------------------------
        #     # fp_lp = 15                                                       # 通過域端周波数[Hz]
        #     # fs_lp = 30                                                       # 阻止域端周波数[Hz]

        #     fp_lp = 3                                                      # 通過域端周波数[Hz]
        #     fs_lp = 5                                                       # 阻止域端周波数[Hz]

        #     gpass = 3                                                        # 通過域端最大損失[dB]
        #     gstop = 40                                                       # 阻止域端最小損失[dB]

        #     # 列方向に順次フィルタ処理をするコード
        #     for i in range(len(df.T)-1):
        #         data = df.T.iloc[i+1]                       # フィルタ処理するデータ列を抽出

        #         # フィルタ処理の種類を文字列で読み取って適切な関数を選択する
        #         if type == 'lp':
        #             # ローパスフィルタを実行
        #             # print('wave=', i, ':Lowpass.')
        #             data = lowpass(x=data, samplerate=1 / dt,
        #                         fp=fp_lp, fs=fs_lp,
        #                         gpass=gpass, gstop=gstop)
        #         else:
        #             # 文字列が当てはまらない時はパス（動作テストでフィルタかけたくない時はNoneとか書いて実行するとよい）
        #             pass

        #         data = pd.Series(data)                                       # dataをPandasシリーズデータへ変換
        #         df_filter[df.columns[i + 1] + '_filter'] = data              # 保存用にデータフレームへdataを追加

        #     df_filter.to_csv(out_file)                                       # フィルタ処理の結果をcsvに保存

        #     return df, df_filter

        # # p=filename
        # # p=str(p.split('.')[0])
        # # print(p)
        # df, df_filter = csv_filter(in_file='C:/Users/BRLAB/Desktop/jushin/'+str(file_cope)+'.csv', out_file='C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv', type='lp')
        # # ws.cell(val, 1).value = p
        # #print(ws.cell(1, 2).value) (1,b)
        # with open('C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv', 'r') as csvfile:
        # # with open('101_1_0_20211204.csv', 'r') as csvfile:
        #     # print(p)
        #     reader = csv.reader(csvfile)
        #     csv_data = list(reader)
        #     # sum=((int(csv_data[1][1])-int(csv_data[2][1]))**2+(int(csv_data[1][2])-int(csv_data[2][2]))**2)**0.5
        #     # print(sum)
        #     # print(csv_data[1][1])
        #     # print(csv_data[1][2])
        #     sum_all=0
        #     sum_all=float(sum_all)
        #     sum_x=0
        #     sum_x=float(sum_x)
        #     sum_y=0
        #     sum_y=float(sum_y)
        #     df = pd.read_csv('C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv',index_col=0)
        #     max_x=df.max()['xposition_filter']
        #     min_x=df.min()['xposition_filter']
        #     maxd_x=max_x-min_x
        #     max_y=df.max()['yposition_filter']
        #     min_y=df.min()['yposition_filter']
        #     maxd_y=max_y-min_y
        #     for i in range(2400):
        #         sum_all=sum_all+((float(csv_data[i+1][2])-float(csv_data[i+2][2]))**2+(float(csv_data[i+1][3])-float(csv_data[i+2][3]))**2)**0.5
        #         sum_x=sum_x+abs(float(csv_data[i+1][2])-float(csv_data[i+2][2]))
        #         sum_y=sum_y+abs(float(csv_data[i+1][3])-float(csv_data[i+2][3]))
        #     # print(sum_all)
        # # ws.cell(val, 6).value = str(sum_all)
        # # ws.cell(val, 7).value = str(sum_x)
        # # ws.cell(val, 8).value = str(sum_y)
        # # ws.cell(val, 9).value = str(maxd_x)
        # # ws.cell(val, 10).value = str(maxd_y)
        # #print(ws.cell(1, 2).value) (1,b)
        # # wb.save('jushin_kekka.xlsx')
        # val=val+1

        # fig = plt.figure(figsize=[15,15])

        # input_csv = pd.read_csv('C:/Users/BRLAB/Desktop/jushin/csv_filter/'+str(file_cope)+'_filter.csv')
        # first_column_data = input_csv[input_csv.keys()[2]]
        # second_column_data = input_csv[input_csv.keys()[3]]

        # plt.xlabel(input_csv.keys()[2],fontsize=20)
        # plt.ylabel(input_csv.keys()[3],fontsize=20)

        # plt.plot(first_column_data, second_column_data, linestyle='solid', marker='')
        # plt.tick_params(labelsize=20)
        # #plt.show()
        # fig.savefig('C:/Users/BRLAB/Desktop/jushin/csv_filter_photo/'+str(file_cope)+"_filter.png")


        # print("総軌跡長＝",sum_all)
        # print("x最大変位=",maxd_x)
        # print("y最大変位＝",maxd_y)

        # xy_area = maxd_x*maxd_y
        # print("矩形面積＝",xy_area)

        # ws = wb.active
        # enumber = 1
        # while not ws.cell(enumber, 1).value is None:
        #     enumber += 1
        # enumber -= 1
        # ws.cell(enumber, 9).value = float(sum_all)
        # ws.cell(enumber, 10).value = float(maxd_x)
        # ws.cell(enumber, 11).value = float(maxd_y)
        # ws.cell(enumber, 12).value = float(xy_area)


        # wb.save('/Users/BRLAB/hikensya.xlsx')

    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_jushin, text="重心を35秒間測定します\n直立を維持してください\n開始まで",font=("",30))
    label2.place(x=30, y=0)
    label2['fg']="black"

    label = tk.Label(frame_jushin, text='0', background='lightgray',font=("",50))
    label.place(x=50, y=130, width=400, height=105)

    thread3 = threading.Thread(target=timer)
    thread3.start()
# 重心動揺測定終了後、「はじめに戻る」もしくは「画像処理/解析」のボタン表示 ###########################################################
def step7_finish_sokutei(): # 測定終了画面
    frame_fini = tk.Frame(baseGround, width=500, height=600) # 測定終了画面、フレーム、
    frame_fini.pack() # 配置、特に指定なし
    def click_btnStart1():   # buttonStart1 が押されたときのイベント
        global jnum
        jnum=1
        frame_fini.destroy()  # 測定終了画面を消す
        step1_start()        # start_jouken を実行

    def click_btnStart2(): # buttonStart2 が押されたときのイベント
        frame_fini.destroy()  # 測定終了画面を消す
        kaiseki_gui()         # kaiseki_guiを実行

    def click_btnStart3(): # buttonStart3 が押されたときのイベント
        frame_fini.destroy()  # 測定終了画面を消す
        shindan1()            # shindan1 を実行
    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_fini, text="計測終了です。",font=("",30))
    label2.place(x=120, y=100)
    label2['fg']="black"

    buttonStart1 = tk.Button(frame_fini,text = 'はじめに戻る', command=click_btnStart1,font=("",20))
    buttonStart1.place(x=50, y=195, relwidth=0.75, relheight= 0.125)

    buttonStart2 = tk.Button(frame_fini,text = '画像処理/解析', command=click_btnStart2,font=("",20))
    buttonStart2.place(x=50, y=295, relwidth=0.75, relheight= 0.125)

    buttonStart3 = tk.Button(frame_fini,text = '診断', command=click_btnStart3,font=("",20))
    buttonStart3.place(x=50, y=395, relwidth=0.75, relheight= 0.125)

# ここから「診断2 」入力 #####################################################################################
def shindan2():
    def click_btn1(): # 'はい' が押されたとき
        if label['text'] == 'いいえ':
            label['text'] = 'はい'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        elif label['text'] == '不明':
            label['text'] = 'はい'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = 'はい'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btn2(): # 'いいえ' が押されたとき
        if label['text'] == 'はい':
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        elif label['text'] == '不明':
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btn3(): # '不明' が押されたとき
        if label['text'] == 'はい':
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        elif label['text'] == 'いいえ':
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btnFinish(): # 「中断」ボタンを押されたときのイベント（スタート画面へ）
        frame_shindan2.destroy()
        step7_finish_sokutei()
    def click_btnEnter(): # 「決定」ボタンが押されたときのイベント（はい/いいえ/不明を番号でエクセルに記録、その後shindan2()へ）
        global komoku2
        komoku2=0
        if label['text'] == 'はい':
            komoku2=1
            # ws = wb.active # エクセルをアクティブに
            # ws.cell(enumber, 14).value = komoku2 # 指定のセルに書き込み
            # wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan2.destroy()
            step7_finish_sokutei()  # 測定終了画面へ、診断項目を増やしたければ、上に新たなshindan3 でも作って、ここをshindan3 と書けばそこに飛ぶ
        elif label['text'] == 'いいえ':
            komoku2=2
            # ws = wb.active # エクセルをアクティブに
            # ws.cell(enumber, 14).value = komoku2 # 指定のセルに書き込み
            # wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan2.destroy()
            step7_finish_sokutei()  # 測定終了画面へ、診断項目を増やしたければ、上に新たなshindan3 でも作って、ここをshindan3 と書けばそこに飛ぶ
        elif label['text'] == '不明':
            komoku2=3
            # ws = wb.active # エクセルをアクティブに
            # ws.cell(enumber, 14).value = komoku2 # 指定のセルに書き込み
            # wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan2.destroy()
            step7_finish_sokutei() # 測定終了画面へ、診断項目を増やしたければ、上に新たなshindan3 でも作って、ここをshindan3 と書けばそこに飛ぶ
        else:
            label2['fg']="red"
    frame_shindan2 = tk.Frame(baseGround, width=600, height=600)
    frame_shindan2.pack()
    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_shindan2, text="現在、糖尿病神経障害罹患者ですか？",font=("",20))
    label2.place(x=50, y=0) # 配置
    label = tk.Label(frame_shindan2, text='', background='lightgray',font=("",30))
    label.place(x=100, y=50, width=400, height=105) # 配置
    button1 = tk.Button(frame_shindan2,text = 'はい', command=click_btn1,font=("",20))
    button1.place(x=50, y=205, relwidth=0.16, relheight= 0.125) # 配置
    button2 = tk.Button(frame_shindan2,text = 'いいえ', command=click_btn2,font=("",20))
    button2.place(x=153, y=205, relwidth=0.16, relheight= 0.125) # 配置
    button3 = tk.Button(frame_shindan2,text = '不明', command=click_btn3,font=("",20))
    button3.place(x=256, y=205, relwidth=0.16, relheight= 0.125) # 配置
    buttonEnter = tk.Button(frame_shindan2,text = '決定', command=click_btnEnter,font=("",30))
    buttonEnter.place(x=359, y=205, relwidth=0.32, relheight= 0.125) # 配置
    buttonFinish = tk.Button(frame_shindan2,text = '中断', command=click_btnFinish)
    buttonFinish.place(x=455, y=295, relwidth=0.16, relheight= 0.125) # 配置

# ここから「診断１」入力 #####################################################################################
def shindan1():
    def click_btn1(): # 'はい' が押されたとき
        if label['text'] == 'いいえ':  # 'はい' が押されたときにlabelを"はい"にする。
            label['text'] = 'はい'     # label は"はい" "いいえ" "不明" の返答結果を表示するラベル
            buttonEnter['fg'] = "red" # buttonEnter は「決定」ボタン
            label2['fg']="black"      # label2 は「現在、糖尿病罹患者ですか？」の文のラベル
        elif label['text'] == '不明':
            label['text'] = 'はい'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = 'はい'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btn2(): # 'いいえ' が押されたとき
        if label['text'] == 'はい': # 'いいえ' が押されたときにlabelを"いいえ"にする。
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        elif label['text'] == '不明':
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = 'いいえ'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btn3(): # '不明' が押されたとき
        if label['text'] == 'はい': # '不明' が押されたときにlabelを"不明"にする。
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        elif label['text'] == 'いいえ':
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
        else:
            label['text'] = '不明'
            buttonEnter['fg'] = "red"
            label2['fg']="black"
    def click_btnFinish(): # 「中断」ボタンを押されたときのイベント（スタート画面へ戻る）
        frame_shindan1.destroy() # frame_shindan1 画面を消す
        step7_finish_sokutei() # 1794行目あたりの finish_sokutei を実行
    def click_btnEnter(): # 「決定」ボタンが押されたときのイベント（はい/いいえ/不明を番号でエクセルに記録、その後shindan2()へ）
        global komoku1
        komoku1=0 # 項目1 の結果をエクセルに送る用の番号
        if label['text'] == 'はい':
            komoku1=1
            # ws = wb.active # 記録用エクセルファイルをアクティブに
            # ws.cell(enumber, 13).value = komoku1 # 指定セルに書き込み
            # wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan1.destroy() # frame_shindan1 画面を消す
            shindan2() # 診断2 へ
        elif label['text'] == 'いいえ':
            komoku1=2
            # ws = wb.active # エクセルをアクティブに
            # ws.cell(enumber, 13).value = komoku1 # 指定セルに書き込み
            # wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan1.destroy()
            shindan2() # 診断2 へ
        elif label['text'] == '不明':
            komoku1=3
            # ws = wb.active # エクセルをアクティブに
            # ws.cell(enumber, 13).value = komoku1 # 指定セルに書き込み
            # wb.save('/Users/BRLAB/hikensya.xlsx') # エクセルを保存
            frame_shindan1.destroy()
            shindan2() # 診断2 へ
        else:
            label2['fg']="red"
    frame_shindan1 = tk.Frame(baseGround, width=600, height=600)
    frame_shindan1.pack()
    # ラベルやボタンの作成と配置
    label2 = tk.Label(frame_shindan1, text="現在、糖尿病罹患者ですか？",font=("",20))
    label2.place(x=50, y=0) # 配置
    label = tk.Label(frame_shindan1, text='', background='lightgray',font=("",30))
    label.place(x=100, y=50, width=400, height=105) # 配置
    button1 = tk.Button(frame_shindan1,text = 'はい', command=click_btn1,font=("",20))
    button1.place(x=50, y=205, relwidth=0.16, relheight= 0.125) # 配置
    button2 = tk.Button(frame_shindan1,text = 'いいえ', command=click_btn2,font=("",20))
    button2.place(x=153, y=205, relwidth=0.16, relheight= 0.125) # 配置
    button3 = tk.Button(frame_shindan1,text = '不明', command=click_btn3,font=("",20))
    button3.place(x=256, y=205, relwidth=0.16, relheight= 0.125) # 配置
    buttonEnter = tk.Button(frame_shindan1,text = '決定', command=click_btnEnter,font=("",30))
    buttonEnter.place(x=359, y=205, relwidth=0.32, relheight= 0.125) # 配置
    buttonFinish = tk.Button(frame_shindan1,text = '中断', command=click_btnFinish)
    buttonFinish.place(x=455, y=295, relwidth=0.16, relheight= 0.125) # 配置

# メインウィンドウを作成
baseGround = tk.Tk()
# ウィンドウのサイズを設定
# baseGround.geometry('500x600')
baseGround.state('zoomed')
# 画面タイトル
baseGround.title('足裏測定')
step1_start()

baseGround.mainloop()